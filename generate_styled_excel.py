"""
生成带样式的选号示例Excel文件
- 命中号码：红色字体 + 黄色背景
- 使用方法：先在网页点击"导出Excel"生成sample_data.json，再运行此脚本
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_FILE = os.path.join(os.path.dirname(__file__), "sample_data.json")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "选号示例_带样式.xlsx")

# 样式定义
HIT_FONT = Font(color="FF0000", bold=True, size=11)
HIT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DRAW_FONT = Font(bold=True, size=11, color="0000FF")
NORMAL_FONT = Font(size=11)
CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def main():
    if not os.path.exists(DATA_FILE):
        print(f"错误：找不到数据文件 {DATA_FILE}")
        print("请先在网页中点击「导出Excel」按钮，再运行此脚本。")
        return

    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    draw_front = set(data["drawFront"])
    draw_back = set(data["drawBack"])
    groups = data["groups"]

    wb = Workbook()
    ws = wb.active
    ws.title = "示例对比"

    # 表头
    headers = ["组别", "前区1", "前区2", "前区3", "前区4", "前区5",
               "后区1", "后区2", "命中前区", "命中后区", "总命中"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 开奖号码行
    ws.cell(row=2, column=1, value="开奖号码").font = DRAW_FONT
    ws.cell(row=2, column=1).alignment = CENTER
    ws.cell(row=2, column=1).border = THIN_BORDER
    for j, num in enumerate(data["drawFront"][:5]):
        cell = ws.cell(row=2, column=j + 2, value=num)
        cell.font = DRAW_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    for j, num in enumerate(data["drawBack"][:2]):
        cell = ws.cell(row=2, column=j + 7, value=num)
        cell.font = DRAW_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    for col in range(9, 12):
        cell = ws.cell(row=2, column=col, value="")
        cell.border = THIN_BORDER

    # 前五组数据
    for i, group in enumerate(groups):
        row = i + 3
        front = group["front"][:5]
        back = group["back"][:2]

        # 组别列
        cell = ws.cell(row=row, column=1, value=group["label"])
        cell.font = NORMAL_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

        # 前区号码 - 命中用红字黄底
        front_hits = 0
        for j, num in enumerate(front):
            cell = ws.cell(row=row, column=j + 2, value=num)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            if num in draw_front:
                cell.font = HIT_FONT
                cell.fill = HIT_FILL
                front_hits += 1
            else:
                cell.font = NORMAL_FONT

        # 后区号码 - 命中用红字黄底
        back_hits = 0
        for j, num in enumerate(back):
            cell = ws.cell(row=row, column=j + 7, value=num)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            if num in draw_back:
                cell.font = HIT_FONT
                cell.fill = HIT_FILL
                back_hits += 1
            else:
                cell.font = NORMAL_FONT

        # 命中统计
        cell = ws.cell(row=row, column=9, value=front_hits)
        cell.font = Font(bold=True, color="FF0000" if front_hits > 0 else "000000", size=11)
        cell.alignment = CENTER
        cell.border = THIN_BORDER

        cell = ws.cell(row=row, column=10, value=back_hits)
        cell.font = Font(bold=True, color="FF0000" if back_hits > 0 else "000000", size=11)
        cell.alignment = CENTER
        cell.border = THIN_BORDER

        total = front_hits + back_hits
        cell = ws.cell(row=row, column=11, value=total)
        cell.font = Font(bold=True, color="FF0000" if total > 0 else "000000", size=11)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if total > 0:
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # 候选池号码
    pool = data.get("poolNumbers", [])
    if pool:
        pool_row = len(groups) + 4
        ws.cell(row=pool_row, column=1, value="候选池").font = Font(bold=True, size=11)
        ws.cell(row=pool_row, column=1).alignment = CENTER
        for j, num in enumerate(pool[:25]):
            cell = ws.cell(row=pool_row, column=j + 2, value=num)
            cell.font = NORMAL_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            # 池中命中号码也标红
            if num in draw_front:
                cell.font = HIT_FONT
                cell.fill = HIT_FILL

    # 设置列宽
    col_widths = [10, 8, 8, 8, 8, 8, 8, 8, 10, 10, 10]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[chr(65 + i)].width = w

    wb.save(OUTPUT_FILE)
    print(f"样式版Excel已生成：{OUTPUT_FILE}")
    print(f"开奖号码：前区 {sorted(data['drawFront'])} 后区 {sorted(data['drawBack'])}")
    for g in groups:
        fh = len([n for n in g["front"][:5] if n in draw_front])
        bh = len([n for n in g["back"][:2] if n in draw_back])
        print(f"  {g['label']}: 命中前区{fh} 后区{bh} 共{fh+bh}球")

if __name__ == "__main__":
    main()
