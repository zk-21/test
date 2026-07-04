import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 读取JSON
with open('hit_rate_detail.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

summary = data['summary']
per_period = data['perPeriod']

# 创建Excel工作簿
wb = Workbook()

# 汇总工作表
ws_summary = wb.active
ws_summary.title = '汇总'

# 写入汇总数据
ws_summary.append(['指标', '值', '命中数', '总球数', '命中率'])
for key, value in summary.items():
    if isinstance(value, dict):
        total_hits = value.get('totalHits', '')
        total_balls = value.get('totalBalls', '')
        hit_rate = value.get('hitRate', '')
        ws_summary.append([key, '', total_hits, total_balls, hit_rate])
    else:
        ws_summary.append([key, value, '', '', ''])

# 格式化汇总表头
for cell in ws_summary[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill('solid', fgColor='D9E1F2')
    cell.alignment = Alignment(horizontal='center')

# 详情工作表
ws_detail = wb.create_sheet('每期详情')

# 定义详情表头
headers = [
    '源期号', '目标期号', '号码池命中数', '号码池命中率',
    'Top5联合覆盖', 'Top5+补漏6联合覆盖', '后区命中数',
    'Top1_命中数', 'Top1_号码', 'Top2_命中数', 'Top2_号码',
    'Top3_命中数', 'Top3_号码', 'Top4_命中数', 'Top4_号码',
    'Top5_命中数', 'Top5_号码', '补漏6_命中数', '补漏6_号码',
    '目标前区', '目标后区'
]
ws_detail.append(headers)

# 写入每期数据
for period in per_period:
    row = [
        period['sI'],
        period['tI'],
        period['poolHit'],
        period['poolRate'],
        period['top5Union'],
        period['top5b6Union'],
        period['backHit'],
    ]
    # Top1-Top5
    for i in range(5):
        top = period['top5'][i]
        row.append(top['hitCount'])
        row.append(' '.join(map(str, top['numbers'])))
    # 补漏6
    if period['bulou6']:
        row.append(period['bulou6']['hitCount'])
        row.append(' '.join(map(str, period['bulou6']['numbers'])))
    else:
        row.append('')
        row.append('')
    # 目标号码
    row.append(' '.join(map(str, period['targetFront'])))
    row.append(' '.join(map(str, period['targetBack'])))
    ws_detail.append(row)

# 格式化详情表头
for cell in ws_detail[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill('solid', fgColor='D9E1F2')
    cell.alignment = Alignment(horizontal='center')

# 调整列宽
for ws in [ws_summary, ws_detail]:
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

# 保存
output_file = 'hit_rate_analysis.xlsx'
wb.save(output_file)
print(f'Excel文件已生成: {output_file}')