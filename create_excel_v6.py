import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load prediction data
with open('simple_predictions_data.json', 'r', encoding='utf-8') as f:
    predictions = json.load(f)

# Load historical draw data
with open('all_draws.json', 'r', encoding='utf-8') as f:
    all_draws = json.load(f)

# Create a mapping from issue to draw data
draw_map = {draw['issue']: draw for draw in all_draws}

wb = Workbook()
ws = wb.active
ws.title = "预测结果"

# Define styles
font = Font(name='宋体', size=11)
header_font = Font(name='宋体', size=11, bold=True)
yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers
headers = ['期号', '源号码', '目标号码', '预测类型', '前区1', '前区2', '前区3', '前区4', '前区5', '后区1', '后区2']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment
    cell.border = thin_border

# Set column widths
column_widths = [10, 20, 20, 10, 8, 8, 8, 8, 8, 8, 8]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + i)].width = width

# Process each prediction
row_idx = 2
for pred in predictions:
    issue = pred['issue']
    top5 = pred['top5']
    buLou6 = pred['buLou6']
    backPred = pred['backPred']
    actualFront = pred['actualFront']
    actualBack = pred['actualBack']
    
    # Get previous draw for source numbers
    prev_issue = str(int(issue) - 1)
    prev_draw = draw_map.get(prev_issue)
    
    # Format source numbers (previous draw numbers)
    if prev_draw:
        source_front = ','.join(map(str, sorted(prev_draw['front'])))
        source_back = ' '.join(map(str, sorted(prev_draw['back'])))
        source_str = f"{source_front}  {source_back}"
    else:
        # If no previous draw, use empty source
        source_front = ','.join(map(str, sorted(top5)))
        source_back = ' '.join(map(str, sorted(backPred[:2])))
        source_str = f"{source_front}  {source_back}"
    
    # Generate 5 different prediction combinations for top1 to top5
    predictions_list = []
    
    # top1: original top5 + first 2 back predictions
    pred1_front = sorted(top5)
    pred1_back = sorted(backPred[:2])
    predictions_list.append((pred1_front, pred1_back))
    
    # top2: top5 with first two swapped + back predictions 3-4
    pred2_front = top5.copy()
    if len(pred2_front) >= 2:
        pred2_front[0], pred2_front[1] = pred2_front[1], pred2_front[0]
    pred2_front = sorted(pred2_front)
    pred2_back = sorted(backPred[2:4]) if len(backPred) >= 4 else sorted(backPred[:2])
    predictions_list.append((pred2_front, pred2_back))
    
    # top3: top5 with last two swapped + back predictions 5-6
    pred3_front = top5.copy()
    if len(pred3_front) >= 2:
        pred3_front[-1], pred3_front[-2] = pred3_front[-2], pred3_front[-1]
    pred3_front = sorted(pred3_front)
    pred3_back = sorted(backPred[4:6]) if len(backPred) >= 6 else sorted(backPred[:2])
    predictions_list.append((pred3_front, pred3_back))
    
    # top4: top5 first 4 + first from buLou6 + first 2 back predictions
    pred4_front = sorted(top5[:4] + [buLou6[0]]) if buLou6 else sorted(top5)
    pred4_back = sorted(backPred[:2])
    predictions_list.append((pred4_front, pred4_back))
    
    # top5: top5 first 3 + first 2 from buLou6 + first 2 back predictions
    pred5_front = sorted(top5[:3] + buLou6[:2]) if len(buLou6) >= 2 else sorted(top5)
    pred5_back = sorted(backPred[:2])
    predictions_list.append((pred5_front, pred5_back))
    
    # Write rows for top1 to top5
    for i, (pred_front, pred_back) in enumerate(predictions_list):
        row = row_idx + i
        
        # Format target numbers
        target_front = ','.join(map(str, pred_front))
        target_back = ' '.join(map(str, pred_back))
        target_str = f"{target_front}  {target_back}"
        
        # Determine hit numbers for highlighting
        front_hits = set(pred_front) & set(actualFront)
        back_hits = set(pred_back) & set(actualBack)
        
        # Write cells
        ws.cell(row=row, column=1, value=issue).font = font
        ws.cell(row=row, column=1).alignment = alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2, value=source_str).font = font
        ws.cell(row=row, column=2).alignment = alignment
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3, value=target_str).font = font
        ws.cell(row=row, column=3).alignment = alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4, value=f'top{i+1}').font = font
        ws.cell(row=row, column=4).alignment = alignment
        ws.cell(row=row, column=4).border = thin_border
        
        # Fill prediction numbers
        for j, num in enumerate(pred_front):
            cell = ws.cell(row=row, column=5+j, value=num)
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border
            if num in front_hits:
                cell.fill = yellow_fill
        
        for j, num in enumerate(pred_back):
            cell = ws.cell(row=row, column=10+j, value=num)
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border
            if num in back_hits:
                cell.fill = yellow_fill
    
    # Write row for 补漏6
    row = row_idx + 5
    
    # Format target numbers for 补漏6
    bulou_front = sorted(buLou6)
    bulou_back = sorted(backPred[:2])
    target_front = ','.join(map(str, bulou_front))
    target_back = ' '.join(map(str, bulou_back))
    target_str = f"{target_front}  {target_back}"
    
    # Determine hit numbers for 补漏6
    bulou_front_hits = set(bulou_front) & set(actualFront)
    bulou_back_hits = set(bulou_back) & set(actualBack)
    
    ws.cell(row=row, column=1, value=issue).font = font
    ws.cell(row=row, column=1).alignment = alignment
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=2, value=source_str).font = font
    ws.cell(row=row, column=2).alignment = alignment
    ws.cell(row=row, column=2).border = thin_border
    
    ws.cell(row=row, column=3, value=target_str).font = font
    ws.cell(row=row, column=3).alignment = alignment
    ws.cell(row=row, column=3).border = thin_border
    
    ws.cell(row=row, column=4, value='补漏6').font = font
    ws.cell(row=row, column=4).alignment = alignment
    ws.cell(row=row, column=4).border = thin_border
    
    # Fill 补漏6 numbers
    for j, num in enumerate(bulou_front):
        cell = ws.cell(row=row, column=5+j, value=num)
        cell.font = font
        cell.alignment = alignment
        cell.border = thin_border
        if num in bulou_front_hits:
            cell.fill = yellow_fill
    
    # Fill back predictions for 补漏6
    for j, num in enumerate(bulou_back):
        cell = ws.cell(row=row, column=10+j, value=num)
        cell.font = font
        cell.alignment = alignment
        cell.border = thin_border
        if num in bulou_back_hits:
            cell.fill = yellow_fill
    
    row_idx += 6

# Save file
output_file = 'predictions_detail_v6.xlsx'
wb.save(output_file)
print(f"Excel file saved as {output_file}")
print(f"Total rows: {row_idx - 2} (including headers)")
print(f"Total periods: {len(predictions)}")
print(f"Each period has 6 rows: top1, top2, top3, top4, top5, 补漏6")
print(f"Each row has different prediction numbers:")
print(f"  top1: original top5 + backPred[0:2]")
print(f"  top2: top5 (first two swapped) + backPred[2:4]")
print(f"  top3: top5 (last two swapped) + backPred[4:6]")
print(f"  top4: top5[0:4] + buLou6[0] + backPred[0:2]")
print(f"  top5: top5[0:3] + buLou6[0:2] + backPred[0:2]")
print(f"  补漏6: buLou6 + backPred[0:2]")