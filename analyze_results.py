"""
从已生成的predictions_detail_v6.xlsx统计命中率和覆盖率
"""
import openpyxl
import json
from collections import defaultdict

# 加载实际开奖数据
with open('all_draws.json', 'r', encoding='utf-8') as f:
    ALL_DRAWS = json.load(f)
issue_map = {d['issue']: d for d in ALL_DRAWS}

# 加载预测数据
wb = openpyxl.load_workbook('predictions_detail_v6.xlsx')
ws = wb.active

# 解析预测数据
predictions = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    issue = row[0]
    pred_type = row[3]
    front_nums = [row[4], row[5], row[6], row[7], row[8]]
    back_nums = [row[9], row[10]]
    
    if issue not in predictions:
        predictions[issue] = {}
    predictions[issue][pred_type] = {
        'front': set(front_nums),
        'back': set(back_nums)
    }

print(f"解析预测期数: {len(predictions)}")

# 统计指标
stats = {
    'top1_front': [], 'top2_front': [], 'top3_front': [],
    'top4_front': [], 'top5_front': [], 'bulou6_front': [],
    'top1_back': [], 'bulou6_back': [],
}

for issue, preds in predictions.items():
    actual = issue_map.get(issue)
    if not actual:
        continue
    actual_front = set(actual['front'])
    actual_back = set(actual['back'])
    
    for pred_type in ['top1', 'top2', 'top3', 'top4', 'top5']:
        if pred_type in preds:
            hit_front = len(actual_front & preds[pred_type]['front'])
            hit_back = len(actual_back & preds[pred_type]['back'])
            stats[f'{pred_type}_front'].append(hit_front)
            if pred_type == 'top1':
                stats['top1_back'].append(hit_back)
    
    if '补漏6' in preds:
        hit_front = len(actual_front & preds['补漏6']['front'])
        hit_back = len(actual_back & preds['补漏6']['back'])
        stats['bulou6_front'].append(hit_front)
        stats['bulou6_back'].append(hit_back)

# 输出统计
total = len(predictions)
print(f"\n{'='*60}")
print(f"v4.1 引擎回测统计（{total}期）")
print(f"{'='*60}")

for key in ['top1_front', 'top2_front', 'top3_front', 'top4_front', 'top5_front', 'bulou6_front']:
    data = stats[key]
    if not data:
        continue
    avg = sum(data) / len(data)
    dist = defaultdict(int)
    for v in data:
        dist[v] += 1
    
    label = key.replace('_front', '前区').replace('bulou6', '补漏6')
    print(f"\n【{label}命中】（5球 vs 实际5球）")
    print(f"  平均命中: {avg:.2f}/5 ({avg/5*100:.1f}%)")
    for k in sorted(dist.keys()):
        cnt = dist[k]
        pct = cnt / len(data) * 100
        print(f"    命中{k}个: {cnt}期 ({pct:.1f}%)")

# 后区
for key in ['top1_back', 'bulou6_back']:
    data = stats[key]
    if not data:
        continue
    avg = sum(data) / len(data)
    dist = defaultdict(int)
    for v in data:
        dist[v] += 1
    
    label = key.replace('_back', '后区').replace('bulou6', '补漏6')
    print(f"\n【{label}命中】（2球 vs 实际2球）")
    print(f"  平均命中: {avg:.2f}/2 ({avg/2*100:.1f}%)")
    for k in sorted(dist.keys()):
        cnt = dist[k]
        pct = cnt / len(data) * 100
        print(f"    命中{k}个: {cnt}期 ({pct:.1f}%)")

# Top1 + 补漏6 联合覆盖
combined = []
for i in range(len(stats['top1_front'])):
    combined.append(stats['top1_front'][i] + stats['bulou6_front'][i])
avg_combined = sum(combined) / len(combined)
combined_dist = defaultdict(int)
for v in combined:
    combined_dist[v] += 1

print(f"\n【Top1+补漏6联合前区命中】（10球 vs 实际5球）")
print(f"  平均命中: {avg_combined:.2f}/5 ({avg_combined/5*100:.1f}%)")
for k in sorted(combined_dist.keys()):
    cnt = combined_dist[k]
    pct = cnt / len(combined) * 100
    print(f"    命中{k}个: {cnt}期 ({pct:.1f}%)")
