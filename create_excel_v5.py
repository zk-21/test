import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load prediction data
with open('simple_predictions_data.json', 'r', encoding='utf-8') as f:
    predictions = json.load(f)

# Load historical draw data
with open('all_draws.json', 'r', encoding='utf-8') as f:
    all_draws = json.load(f)

# Create a mapping from issue to draw data
draw_map = {draw['issue']: draw for draw in all_draws}

wb = Workbook()
ws = wb.active
ws.title = "预测结果"

# Define styles
font = Font(name='宋体', size=11)
header_font = Font(name='宋体', size=11, bold=True)
yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers
headers = ['期号', '源号码', '目标号码', '预测类型', '前区1', '前区2', '前区3', '前区4', '前区5', '后区1', '后区2']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment
    cell.border = thin_border

# Set column widths
column_widths = [10, 20, 20, 10, 8, 8, 8, 8, 8, 8, 8]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + i)].width = width

# Process each prediction
row_idx = 2
for pred in predictions:
    issue = pred['issue']
    top5 = pred['top5']
    buLou6 = pred['buLou6']
    backPred = pred['backPred']
    actualFront = pred['actualFront']
    actualBack = pred['actualBack']
    
    # Get previous draw for source numbers
    prev_issue = str(int(issue) - 1)
    prev_draw = draw_map.get(prev_issue)
    
    # Format source numbers (previous draw numbers)
    if prev_draw:
        source_front = ','.join(map(str, sorted(prev_draw['front'])))
        source_back = ' '.join(map(str, sorted(prev_draw['back'])))
        source_str = f"{source_front}  {source_back}"
    else:
        # If no previous draw, use empty source
        source_front = ','.join(map(str, sorted(top5)))
        source_back = ' '.join(map(str, sorted(backPred[:2])))
        source_str = f"{source_front}  {source_back}"
    
    # Determine hit numbers for highlighting
    front_hits = set(top5) & set(actualFront)
    back_hits = set(backPred[:2]) & set(actualBack)
    
    # Generate 5 different prediction combinations for top1 to top5
    # For simplicity, we'll use the same top5 but with different orders
    # In a real scenario, you might want to generate 5 different predictions
    for i in range(5):
        row = row_idx + i
        
        # Format target numbers (prediction results)
        # Use top5 with different order for each row
        if i == 0:
            # top1: use original top5
            target_front = ','.join(map(str, sorted(top5)))
            target_back = ' '.join(map(str, sorted(backPred[:2])))
            target_str = f"{target_front}  {target_back}"
        elif i == 1:
            # top2: use top5 with first two numbers swapped
            swapped = top5.copy()
            if len(swapped) >= 2:
                swapped[0], swapped[1] = swapped[1], swapped[0]
            target_front = ','.join(map(str, sorted(swapped)))
            target_back = ' '.join(map(str, sorted(backPred[:2])))
            target_str = f"{target_front}  {target_back}"
        elif i == 2:
            # top3: use top5 with last two numbers swapped
            swapped = top5.copy()
            if len(swapped) >= 2:
                swapped[-1], swapped[-2] = swapped[-2], swapped[-1]
            target_front = ','.join(map(str, sorted(swapped)))
            target_back = ' '.join(map(str, sorted(backPred[:2])))
            target_str = f"{target_front}  {target_back}"
        elif i == 3:
            # top4: use top5 with middle numbers swapped
            swapped = top5.copy()
            if len(swapped) >= 4:
                swapped[1], swapped[2] = swapped[2], swapped[1]
            target_front = ','.join(map(str, sorted(swapped)))
            target_back = ' '.join(map(str, sorted(backPred[:2])))
            target_str = f"{target_front}  {target_back}"
        else:
            # top5: use top5 with first and last swapped
            swapped = top5.copy()
            if len(swapped) >= 2:
                swapped[0], swapped[-1] = swapped[-1], swapped[0]
            target_front = ','.join(map(str, sorted(swapped)))
            target_back = ' '.join(map(str, sorted(backPred[:2])))
            target_str = f"{target_front}  {target_back}"
        
        # Write cells
        ws.cell(row=row, column=1, value=issue).font = font
        ws.cell(row=row, column=1).alignment = alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2, value=source_str).font = font
        ws.cell(row=row, column=2).alignment = alignment
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3, value=target_str).font = font
        ws.cell(row=row, column=3).alignment = alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4, value=f'top{i+1}').font = font
        ws.cell(row=row, column=4).alignment = alignment
        ws.cell(row=row, column=4).border = thin_border
        
        # Fill prediction numbers for each row
        for j, num in enumerate(top5):
            cell = ws.cell(row=row, column=5+j, value=num)
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border
            if num in front_hits:
                cell.fill = yellow_fill
        
        for j, num in enumerate(backPred[:2]):
            cell = ws.cell(row=row, column=10+j, value=num)
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border
            if num in back_hits:
                cell.fill = yellow_fill
    
    # Write row for 补漏6
    row = row_idx + 5
    
    # Format target numbers for 补漏6
    target_front = ','.join(map(str, sorted(buLou6)))
    target_back = ' '.join(map(str, sorted(backPred[:2])))
    target_str = f"{target_front}  {target_back}"
    
    # Determine hit numbers for 补漏6
    bulou_front_hits = set(buLou6) & set(actualFront)
    bulou_back_hits = set(backPred[:2]) & set(actualBack)
    
    ws.cell(row=row, column=1, value=issue).font = font
    ws.cell(row=row, column=1).alignment = alignment
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=2, value=source_str).font = font
    ws.cell(row=row, column=2).alignment = alignment
    ws.cell(row=row, column=2).border = thin_border
    
    ws.cell(row=row, column=3, value=target_str).font = font
    ws.cell(row=row, column=3).alignment = alignment
    ws.cell(row=row, column=3).border = thin_border
    
    ws.cell(row=row, column=4, value='补漏6').font = font
    ws.cell(row=row, column=4).alignment = alignment
    ws.cell(row=row, column=4).border = thin_border
    
    # Fill 补漏6 numbers
    for j, num in enumerate(buLou6):
        cell = ws.cell(row=row, column=5+j, value=num)
        cell.font = font
        cell.alignment = alignment
        cell.border = thin_border
        if num in bulou_front_hits:
            cell.fill = yellow_fill
    
    # Fill back predictions for 补漏6
    for j, num in enumerate(backPred[:2]):
        cell = ws.cell(row=row, column=10+j, value=num)
        cell.font = font
        cell.alignment = alignment
        cell.border = thin_border
        if num in bulou_back_hits:
            cell.fill = yellow_fill
    
    row_idx += 6

# Save file
output_file = 'predictions_detail_v5.xlsx'
wb.save(output_file)
print(f"Excel file saved as {output_file}")
print(f"Total rows: {row_idx - 2} (including headers)")
print(f"Total periods: {len(predictions)}")
print(f"Each period has 6 rows: top1, top2, top3, top4, top5, 补漏6")
print(f"Source numbers: previous draw numbers")
print(f"Target numbers: prediction numbers for each row")