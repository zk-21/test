#!/usr/bin/env python3
"""
v4.1 预测引擎 Python 移植版
基于 optimized_picker.js 的完整预测逻辑
生成 predictions_detail_v6.xlsx 格式的预测文件
"""
import json
import random
import math
from itertools import combinations
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ===================== 加载数据 =====================
with open('all_draws.json', 'r', encoding='utf-8') as f:
    ALL_DRAWS = json.load(f)
print(f"加载 {len(ALL_DRAWS)} 期数据")

issue_map = {d['issue']: d for d in ALL_DRAWS}

# ===================== 配置 =====================
CONFIG = {
    'frontMax': 35, 'backMax': 12, 'poolSize': 24, 'pickCount': 5,
    'offsetScore': {0:20, 1:15, 2:13, 3:12, 4:10, 5:8, 6:6, 7:5, 8:4, 9:3, 10:2},
    'tailSameScore': 35, 'tailNeighborScore': 15, 'tailWithinSource': 5,
    'intervalMatchScore': 30, 'intervalTwoScore': 15,
    'targetSum': 87.5, 'sumTolerance': 17.5, 'targetSpan': 24, 'spanTolerance': 8,
    'extremeSumDrop': 30, 'extremeParityFlip': 4,
    'anchorOverusePenalty': 8, 'extremeZonePenalty': 200, 'sumOutlierPenalty': 15,
    'comboPoolTop': 20, 'comboSampleMax': 500, 'hotBoostWeight': 8,
    'tailPatternBonus': 10, 'comboDiversityMin': 0.3,
    'historyFreqWeight': 0.15, 'recentFreqWeight': 0.10, 'repeatRateWeight': 0.05,
    'prevDrawTailsWeight': 2,
    'runPenaltyScale': 0.3, 'repeatPenaltyScale': 0.8,
    'tailPatternWeight': 0.6, 'diversityScoreWeight': 0.65, 'diversityMinScoreRatio': 0.65,
}

# ===================== 工具函数 =====================
def gi(n):
    if n <= 12: return 0
    if n <= 24: return 1
    return 2

def tail(n): return n % 10

def tails(nums):
    return sorted(set(n % 10 for n in nums))

def sum_nums(nums):
    return sum(nums)

def span(nums):
    s = sorted(nums)
    return s[-1] - s[0]

def odd_count(nums):
    return sum(1 for n in nums if n % 2 == 1)

def interval_ratio(nums):
    iv = [0, 0, 0]
    for n in nums: iv[gi(n)] += 1
    return iv

def iv_key(nums):
    return ':'.join(map(str, interval_ratio(nums)))

def get_interval_ratio_distance(r1, r2):
    return sum(abs(r1[i] - r2[i]) for i in range(3))

def sort_by_score(arr):
    return sorted(arr, key=lambda x: x['score'], reverse=True)

# ===================== 历史频率分析 =====================
def calculate_history_metrics():
    total = len(ALL_DRAWS)
    history_freq = [0]*36
    for d in ALL_DRAWS:
        for n in d['front']: history_freq[n] += 1
    recent_window = 20
    recent_freq = [0]*36
    for d in ALL_DRAWS[-recent_window:]:
        for n in d['front']: recent_freq[n] += 1
    repeat_rate = [0]*36
    for i in range(len(ALL_DRAWS) - 10):
        src = set(ALL_DRAWS[i]['front'])
        tgt = set(ALL_DRAWS[i+10]['front'])
        for n in src:
            if n in tgt: repeat_rate[n] += 1
    repeat_pairs = max(1, len(ALL_DRAWS) - 10)
    norm_repeat = [c / repeat_pairs for c in repeat_rate]
    avg_hist = sum(history_freq) / 35
    avg_recent = sum(recent_freq) / 35
    avg_repeat = sum(norm_repeat) / 35
    return {
        'historyFreq': history_freq, 'recentFreq': recent_freq,
        'normalizedRepeatRate': norm_repeat,
        'avgHistoryFreq': avg_hist, 'avgRecentFreq': avg_recent,
        'avgRepeatRate': avg_repeat, 'totalDraws': total,
    }

history_metrics = calculate_history_metrics()

# ===================== 连号分析 =====================
def build_consecutive_segments(numbers):
    s = sorted(numbers)
    segs, seg = [], [s[0]]
    for i in range(1, len(s)):
        if s[i] - s[i-1] == 1:
            seg.append(s[i])
        else:
            if len(seg) >= 2: segs.append(seg)
            seg = [s[i]]
    if len(seg) >= 2: segs.append(seg)
    return segs

def count_consecutive_pairs(numbers):
    s = sorted(numbers)
    pairs, longest, cur = 0, 1, 1
    for i in range(1, len(s)):
        if s[i] - s[i-1] == 1:
            cur += 1; pairs += 1
            longest = max(longest, cur)
        else:
            cur = 1
    return pairs, longest

def get_run_penalty(numbers, anchor_numbers=[]):
    segs = build_consecutive_segments(numbers)
    anchor_set = set(anchor_numbers)
    longest, penalty, dbl = 0, 0, 0
    for seg in segs:
        longest = max(longest, len(seg))
        support = sum(1 for n in seg if n in anchor_set)
        ratio = support / len(seg) if seg else 0
        disc = 0.45 if ratio >= 0.8 else (0.75 if ratio >= 0.6 else 1)
        if len(seg) == 2:
            dbl += 1; penalty += round(3 * disc)
        elif len(seg) >= 4:
            penalty += round((70 + (len(seg)-4)*16) * disc)
        elif len(seg) == 3:
            penalty += round(36 * disc)
    if dbl >= 3: penalty += (dbl - 2) * 3
    return {'longestRun': longest, 'runPenalty': penalty, 'doubleRunCount': dbl}

def get_repeat_penalty(numbers, source_numbers=[]):
    src_set = set(source_numbers)
    cnt = sum(1 for n in numbers if n in src_set)
    if cnt == 0: p = 0
    elif cnt == 1: p = 4
    elif cnt == 2: p = 10
    elif cnt == 3: p = 30
    else: p = 30 + (cnt-3)*20
    return {'repeatCount': cnt, 'repeatPenalty': p}

def score_tail_patterns(combo_numbers):
    ts = sorted(set(n % 10 for n in combo_numbers))
    score = 0
    longest_c, cur_c = 1, 1
    for i in range(1, len(ts)):
        if ts[i] == ts[i-1] + 1: cur_c += 1; longest_c = max(longest_c, cur_c)
        else: cur_c = 1
    if 0 in ts and 9 in ts:
        wr = 1
        for i in range(len(ts)-1, -1, -1):
            if ts[i] >= 9: wr += 1
            else: break
        longest_c = max(longest_c, wr)
    if longest_c >= 3: score += 40
    elif longest_c >= 2: score += 20
    for d in range(2, 5):
        for start in range(0, 10 - d*2 + 1):
            cnt = 0
            for v in range(start, 10, d):
                if v in ts: cnt += 1
                else: break
            if cnt >= 4: score += 30
            elif cnt >= 3: score += 15
    if len(ts) >= 5: score += 20
    elif len(ts) >= 4: score += 10
    return {'score': score, 'longestConsec': longest_c, 'tailCount': len(ts)}

# ===================== 极端期检测 =====================
def detect_extreme(source_draw, neighbor_draws):
    flags = {'sumCrash': False, 'parityFlip': False, 'narrowRange': False}
    if len(neighbor_draws) >= 2:
        avg_prev = sum(sum_nums(d['front']) for d in neighbor_draws[:2]) / 2
        if abs(sum_nums(source_draw['front']) - avg_prev) > CONFIG['extremeSumDrop']:
            flags['sumCrash'] = True
    if len(neighbor_draws) >= 1:
        if abs(odd_count(source_draw['front']) - odd_count(neighbor_draws[0]['front'])) >= CONFIG['extremeParityFlip']:
            flags['parityFlip'] = True
    if span(source_draw['front']) <= 12:
        flags['narrowRange'] = True
    return flags

# ===================== 热号分析 =====================
def compute_hotness(source_idx, lookback=10):
    freq = defaultdict(int)
    start = max(0, source_idx - lookback)
    for i in range(start, source_idx):
        for n in ALL_DRAWS[i]['front']: freq[n] += 1
    return freq

# ===================== S1: +13期趋势映射 =====================
def build_plus_ten_trend_map(source_idx, lookback=50):
    src = ALL_DRAWS[source_idx]
    src_nums = sorted(src['front'])
    src_tails = set(n % 10 for n in src_nums)
    src_tail_nb = set()
    for t in src_tails:
        src_tail_nb.add(t); src_tail_nb.add((t+1)%10); src_tail_nb.add((t+9)%10)
    src_iv = interval_ratio(src_nums)
    src_iv_k = ':'.join(map(str, src_iv))
    target_map, neighbor_map = defaultdict(float), defaultdict(float)
    end = source_idx - 13
    start = max(0, end - lookback)
    for i in range(start, end + 1):
        hs = ALL_DRAWS[i]; ht = ALL_DRAWS[i+13] if i+13 < len(ALL_DRAWS) else None
        if not hs or not ht: continue
        hn = sorted(hs['front']); h_set = set(hn)
        h_tails = set(n % 10 for n in hn)
        h_tail_nb = set()
        for t in h_tails:
            h_tail_nb.add(t); h_tail_nb.add((t+1)%10); h_tail_nb.add((t+9)%10)
        exact = sum(1 for n in src_nums if n in h_set)
        neighbor = sum(1 for n in src_nums if (n-1) in h_set or (n+1) in h_set)
        tail_ov = sum(1 for n in src_nums if (n%10) in h_tails)
        tail_nb_ov = sum(1 for n in src_nums if (n%10) in h_tail_nb)
        sel_tail_sig = sum(1 for n in hn if (n%10) in src_tails)
        sel_tail_nb_sig = sum(1 for n in hn if (n%10) in src_tail_nb)
        h_iv = interval_ratio(hn)
        ratio_match = 1 if ':'.join(map(str,h_iv)) == src_iv_k else 0
        iv_diff = sum(abs(h_iv[j]-src_iv[j]) for j in range(3))
        iv_sim = max(0, 6 - iv_diff)
        row_dist = abs(i - source_idx)
        prox = 10 if row_dist <= 3 else (6 if row_dist <= 6 else (3 if row_dist <= 10 else 0))
        w = exact*18 + neighbor*10 + tail_ov*8 + tail_nb_ov*4 + sel_tail_sig*5 + sel_tail_nb_sig*2 + ratio_match*16 + iv_sim*3 + prox
        if w <= 0: continue
        for n in ht['front']:
            target_map[n] += w
            for d in range(1, 4):
                for nb in [n-d, n+d]:
                    if 1 <= nb <= CONFIG['frontMax']:
                        neighbor_map[nb] += max(1, round(w * 0.4 * (1 - d*0.2)))
    return {'targetMap': dict(target_map), 'neighborMap': dict(neighbor_map)}

# ===================== S2: 桥梁分析 =====================
def build_bridge_map(anchor_numbers, support_numbers=[]):
    max_gap = 4
    anchors = sorted(anchor_numbers)
    support_set = set(support_numbers)
    support_tail_set = set(n % 10 for n in support_set)
    gap_map, endpoint_map = {}, {}
    for li in range(len(anchors)):
        for ri in range(li+1, len(anchors)):
            left, right = anchors[li], anchors[ri]
            gap = right - left
            if gap <= 1 or gap > max_gap: continue
            closeness = max(1, max_gap - gap + 1)
            for ep in [left, right]:
                cur = endpoint_map.get(ep, {'score':0,'hits':0})
                cur['hits'] += 1; cur['score'] += 8 + closeness*3
                if ep in support_set: cur['score'] += 6
                if (ep-1) in support_set: cur['score'] += 2
                if (ep+1) in support_set: cur['score'] += 2
                endpoint_map[ep] = cur
            for n in range(left+1, right):
                cur = gap_map.get(n, {'score':0,'hits':0})
                cur['hits'] += 1; cur['score'] += 24 + closeness*6
                if n in support_set: cur['score'] += 14
                nb_s = (1 if (n-1) in support_set else 0) + (1 if (n+1) in support_set else 0)
                if nb_s > 0: cur['score'] += nb_s * 4
                if (n%10) in support_tail_set: cur['score'] += 2
                gap_map[n] = cur
    return {'gapMap': gap_map, 'endpointMap': endpoint_map}

# ===================== S3: 等距端点分析 =====================
def build_arithmetic_endpoint_map(anchor_numbers, support_numbers=[], max_gap=6):
    anchors = sorted(anchor_numbers)
    support_set = set(support_numbers)
    endpoint_map = {}
    for a in anchors:
        for d in range(1, max_gap+1):
            for ep in [a-d, a+d]:
                if ep < 1 or ep > CONFIG['frontMax']: continue
                closeness = max(1, max_gap - d + 1)
                cur = endpoint_map.get(ep, {'score':0,'hits':0})
                cur['hits'] += 1; cur['score'] += 10 + closeness*4
                if ep in support_set: cur['score'] += 6
                if (ep-1) in support_set or (ep+1) in support_set: cur['score'] += 2
                endpoint_map[ep] = cur
    return endpoint_map

# ===================== S4: 增强扩散惩罚 =====================
def get_enhanced_spread_penalty(numbers):
    s = sorted(numbers)
    if len(s) <= 1: return {'penalty':0,'span':0,'maxWindowCount':len(s),'coveredIntervals':1}
    sp = s[-1] - s[0]
    iv = [0,0,0]
    for n in s: iv[gi(n)] += 1
    cov_iv = sum(1 for c in iv if c > 0)
    max_iv = max(iv)
    pen = 0; dw = 8
    if cov_iv >= 3:
        if sp <= 18: pen += 2
        if sp <= 16: pen += 6
        if sp <= 13: pen += 10
        if sp <= 10: pen += 16
    elif cov_iv == 2:
        if sp <= 12: pen += 3
        if sp <= 10: pen += 7
        if sp <= 8: pen += 12
        if sp <= 6: pen += 16
    else:
        if sp <= 7: pen += 2
        if sp <= 5: pen += 6
        if sp <= 3: pen += 10
    max_wc = 0
    for si in range(len(s)):
        ei = si
        while ei < len(s) and s[ei] - s[si] <= dw: ei += 1
        cnt = ei - si; max_wc = max(max_wc, cnt)
        if cov_iv >= 3:
            if cnt >= 4: pen += 14 + (cnt-4)*8
            elif cnt == 3: pen += 4
        elif cov_iv == 2:
            if cnt >= 4: pen += 10 + (cnt-4)*6
        else:
            if cnt >= 4: pen += 8 + (cnt-4)*4
    if cov_iv >= 3 and max_iv >= 4: pen += 10 + (max_iv-4)*6
    elif cov_iv == 2 and max_iv >= 4: pen += 8 + (max_iv-4)*4
    return {'penalty':pen,'span':sp,'maxWindowCount':max_wc,'coveredIntervals':cov_iv,'maxIntervalCount':max_iv}

# ===================== S5: 参考行分析 =====================
def build_reference_window(source_idx, lookback=5):
    refs = []
    start = max(0, source_idx - lookback)
    for i in range(start, source_idx):
        d = ALL_DRAWS[i]
        if not d: continue
        nums = sorted(d['front'])
        ts = set(n % 10 for n in nums)
        iv = interval_ratio(nums)
        cp, lr = count_consecutive_pairs(nums)
        cs = build_consecutive_segments(nums)
        ae = set()
        for diff in range(2, 7):
            for n in nums:
                a, b = n - diff, n + diff
                if 1 <= a <= 35 and a in set(nums): ae.add(n); ae.add(a)
                if 1 <= b <= 35 and b in set(nums): ae.add(n); ae.add(b)
        bg, be = set(), set()
        for j in range(len(nums)-1):
            g = nums[j+1] - nums[j]
            if 2 <= g <= 4:
                for m in range(nums[j]+1, nums[j+1]): bg.add(m)
                be.add(nums[j]); be.add(nums[j+1])
        tc = Counter(n % 10 for n in nums)
        st, sc = None, 0
        for t, c in tc.items():
            if c > sc: sc = c; st = t
        refs.append({
            'row': i, 'numbers': nums, 'numberSet': set(nums), 'tailSet': ts,
            'ivKey': ':'.join(map(str,iv)), 'iv': iv,
            'consecutivePairs': cp, 'longestRun': lr, 'consecutiveSegments': cs,
            'arithEndpoints': ae, 'bridgeGaps': bg, 'bridgeEndpoints': be,
            'strongestTail': st, 'strongestCount': sc,
        })
    return refs

# ===================== S6: 组合vs参考行评分 =====================
def score_combo_against_references(combo_numbers, refs):
    total, satisfied = 0, 0
    for ref in refs:
        rs, signals = 0, 0
        to = sum(1 for n in combo_numbers if (n%10) in ref['tailSet'])
        rs += min(to, 3) * 8
        if to >= 1: signals += 1
        tnb = set()
        for t in ref['tailSet']: tnb.add((t+1)%10); tnb.add((t+9)%10)
        tn = sum(1 for n in combo_numbers if (n%10) in tnb)
        rs += min(tn, 3) * 4
        if tn >= 1: signals += 1
        ov = sum(1 for n in combo_numbers if n in ref['numberSet'])
        rs += min(ov, 3) * 8
        if ov >= 1: signals += 1
        nbh = sum(1 for n in combo_numbers if (n-1) in ref['numberSet'] or (n+1) in ref['numberSet'])
        rs += min(nbh, 3) * 4
        if nbh >= 1: signals += 1
        civ = ':'.join(map(str, interval_ratio(combo_numbers)))
        if civ == ref['ivKey']: rs += 12; signals += 1
        if ref['strongestCount'] >= 2 and ref['strongestTail'] is not None:
            sh = sum(1 for n in combo_numbers if n%10 == ref['strongestTail'])
            if sh >= 1: signals += 1
            rs += sh * 6
        if ref.get('arithEndpoints') and len(ref['arithEndpoints']) > 0:
            ah = sum(1 for n in combo_numbers if n in ref['arithEndpoints'])
            rs += ah * 5
            if ah >= 2: rs += 6; signals += 1
        if ref.get('bridgeGaps') and len(ref['bridgeGaps']) > 0:
            gh = sum(1 for n in combo_numbers if n in ref['bridgeGaps'])
            eh = sum(1 for n in combo_numbers if n in ref['bridgeEndpoints'])
            rs += gh*6 + eh*4
            if gh+eh >= 2: rs += 5; signals += 1
        cp2, lr2 = count_consecutive_pairs(combo_numbers)
        ps = max(0, 3 - abs(cp2 - ref['consecutivePairs'])) if ref['consecutivePairs'] > 0 else (1 if cp2 == 0 else 0)
        rns = max(0, 3 - abs(lr2 - ref['longestRun'])) if ref['longestRun'] > 1 else (1 if lr2 <= 2 else 0)
        rs += (ps + rns) * 3
        if ps >= 2 or rns >= 2: signals += 1
        for seg in ref.get('consecutiveSegments', []):
            seg_set = set(seg)
            shared = sum(1 for n in combo_numbers if n in seg_set)
            adj = sum(1 for n in combo_numbers if (n-1) in seg_set or (n+1) in seg_set)
            if shared >= min(2, len(seg)): rs += 8; signals += 1
            elif adj > 0: rs += 3
        if rs >= 20: satisfied += 1
        total += rs
    return {'score': total, 'satisfiedRows': satisfied}

# ===================== 尾号转移分析 =====================
def analyze_tail_transitions(source_idx, lookback=12):
    trans_freq = defaultdict(int)
    tail_freq = defaultdict(int)
    start = max(0, source_idx - lookback)
    for i in range(start, source_idx - 10):
        src = ALL_DRAWS[i]; tgt = ALL_DRAWS[i+10] if i+10 < len(ALL_DRAWS) else None
        if not src or not tgt: continue
        st = tails(src['front']); tt = tails(tgt['front'])
        for t in tt: tail_freq[t] += 1
        for s in st:
            for t in tt: trans_freq[f"{s}→{t}"] += 1
    return {'transFreq': dict(trans_freq), 'tailFreq': dict(tail_freq)}

def predict_likely_tails(source_tails, trans_data):
    """基础版尾号预测"""
    scores = defaultdict(float)
    for st in source_tails:
        for tt in range(10):
            key = f"{st}→{tt}"
            scores[tt] += trans_data['transFreq'].get(key, 0)
    for t, cnt in trans_data['tailFreq'].items():
        scores[t] += cnt * 0.5
    return sorted(scores.items(), key=lambda x: -x[1])

def get_arithmetic_tails(tail_set):
    """等差延伸尾号计算"""
    arith_tails = set()
    sorted_tails = sorted(tail_set)
    for i in range(len(sorted_tails)):
        for j in range(i + 1, len(sorted_tails)):
            diff = sorted_tails[j] - sorted_tails[i]
            if diff in [2, 3, 4]:
                arith_tails.add(sorted_tails[i])
                arith_tails.add(sorted_tails[j])
                # 延伸
                next_t = (sorted_tails[j] + diff) % 10
                prev_t = (sorted_tails[i] - diff + 10) % 10
                arith_tails.add(next_t)
                arith_tails.add(prev_t)
    return arith_tails

def predict_likely_tails_enhanced(source_tails, trans_data, source_idx, lookback=50):
    """增强版尾号预测（融合全局高频+等差延伸+参考行重叠）"""
    weights = {
        'overlap1': 6,       # 上期尾号重叠
        'arith1': 12,        # 上期等差延伸
        'overlap10': 4,      # 上10期尾号重叠
        'arith10': 10,       # 上10期等差延伸
        'overlapBonus': 2,   # 双期重叠奖励
        'globalFreq': 28,    # 全局高频权重
    }
    
    scores = defaultdict(float)
    
    # 1. 原始转移概率
    for st in source_tails:
        for tt in range(10):
            key = f"{st}→{tt}"
            scores[tt] += trans_data['transFreq'].get(key, 0)
    
    # 2. 全局高频尾号（近50期）
    global_tail_freq = defaultdict(int)
    actual_lookback = min(lookback, source_idx)
    for i in range(max(0, source_idx - actual_lookback), source_idx):
        for t in tails(ALL_DRAWS[i]['front']):
            global_tail_freq[t] += 1
    max_global = max(1, max(global_tail_freq.values()) if global_tail_freq else 1)
    for t, cnt in global_tail_freq.items():
        scores[t] += (cnt / max_global) * weights['globalFreq']
    
    # 3. 参考行尾号重叠（上一期 = source_idx - 1）
    if source_idx >= 1:
        ref1_tails = set(tails(ALL_DRAWS[source_idx - 1]['front']))
        for t in ref1_tails:
            scores[t] += weights['overlap1']
        # 等差延伸
        arith1 = get_arithmetic_tails(ref1_tails)
        for t in arith1:
            if t not in ref1_tails:
                scores[t] += weights['arith1']
    
    # 4. 参考行尾号重叠（上10期 = source_idx - 10）
    if source_idx >= 10:
        ref10_tails = set(tails(ALL_DRAWS[source_idx - 10]['front']))
        for t in ref10_tails:
            scores[t] += weights['overlap10']
        # 等差延伸
        arith10 = get_arithmetic_tails(ref10_tails)
        for t in arith10:
            if t not in ref10_tails:
                scores[t] += weights['arith10']
        
        # 5. 双期重叠奖励（上期和上10期都有的尾号）
        if source_idx >= 1:
            overlap = ref1_tails & ref10_tails
            for t in overlap:
                scores[t] += weights['overlapBonus']
    
    return sorted(scores.items(), key=lambda x: -x[1])

# ===================== 区间比预测 =====================
def predict_target_interval_ratio(source_idx, source_iv):
    src_key = ':'.join(map(str, source_iv))
    transitions = defaultdict(lambda: {'count':0, 'weight':0})
    window_size = min(60, source_idx)
    specific_count = 0
    global_dist_sum, global_dist_cnt = 0, 0
    for i in range(source_idx):
        if i >= len(ALL_DRAWS) - 1: continue
        s_iv = interval_ratio(ALL_DRAWS[i]['front'])
        t_iv = interval_ratio(ALL_DRAWS[i+1]['front'])
        global_dist_sum += get_interval_ratio_distance(s_iv, t_iv)
        global_dist_cnt += 1
        if ':'.join(map(str,s_iv)) != src_key: continue
        specific_count += 1
        tk = ':'.join(map(str,t_iv))
        recency = 1 + (i - max(0, source_idx - window_size)) / window_size * 2
        transitions[tk]['count'] += 1
        transitions[tk]['weight'] += recency
    global_avg = global_dist_sum / max(1, global_dist_cnt)
    min_specific = 4
    blend = min(1, specific_count / min_specific)
    if not transitions:
        return {'predictedIv': source_iv, 'predictedIvKey': src_key, 'distance': round(global_avg), 'confidence': 0, 'topCandidates': []}
    max_w = max(1, max(d['weight'] for d in transitions.values()))
    sorted_t = sorted(transitions.items(), key=lambda x: -(x[1]['count']*0.7 + (x[1]['weight']/max_w)*30))
    top3 = sorted_t[:3]
    pred_iv = [int(x) for x in top3[0][0].split(':')]
    raw_dist = get_interval_ratio_distance(source_iv, pred_iv)
    total_score = sum(d['count']*0.7 + (d['weight']/max_w)*30 for _, d in top3)
    raw_conf = (top3[0][1]['count']*0.7 + (top3[0][1]['weight']/max_w)*30) / max(0.1, total_score)
    blended_dist = round(raw_dist * blend + global_avg * (1-blend))
    conf = raw_conf * blend
    top_cands = [{'iv': [int(x) for x in k.split(':')], 'ivKey': k, 'score': d['count']*0.7+(d['weight']/max_w)*30, 'count': d['count']} for k, d in top3]
    return {'predictedIv': pred_iv, 'predictedIvKey': top3[0][0], 'distance': blended_dist, 'confidence': conf, 'topCandidates': top_cands}

# ===================== 候选号码池生成 =====================
def generate_candidate_pool(source_draw, target_tails, target_iv, extreme_flags, hotness, extra_maps, iv_prediction, target_draw, predicted_tails):
    anchors = source_draw['front']
    src_tails = tails(anchors)
    src_iv = interval_ratio(anchors)
    src_odd = odd_count(anchors)
    src_sum = sum_nums(anchors)
    tgt_odd = odd_count(target_draw['front']) if target_draw else None
    tgt_sum = sum_nums(target_draw['front']) if target_draw else None
    pt_map = extra_maps.get('plusTenTargetMap', {})
    pn_map = extra_maps.get('plusTenNeighborMap', {})
    bg_map = extra_maps.get('bridgeGapMap', {})
    be_map = extra_maps.get('bridgeEndpointMap', {})
    ae_map = extra_maps.get('arithmeticEndpointMap', {})
    prev_tails = extra_maps.get('prevDrawTails', [])
    max_pt = max(1, max(pt_map.values())) if pt_map else 1
    max_br = max(1, max((v['score'] for v in bg_map.values()), default=1), max((v['score'] for v in be_map.values()), default=1))
    max_ar = max(1, max((v['score'] for v in ae_map.values()), default=1))
    candidates = []
    for n in range(1, CONFIG['frontMax']+1):
        score = 0
        min_off = 999
        best_anc = None
        for a in anchors:
            dist = abs(n - a)
            if dist < min_off: min_off = dist; best_anc = a
        score += CONFIG['offsetScore'].get(min_off, 0)
        t = tail(n)
        if t in src_tails: score += 15
        if prev_tails and t in prev_tails: score += 8
        if target_tails and t in target_tails: score += 35
        elif target_tails and any(abs(t-tt)==1 for tt in target_tails): score += 15
        if predicted_tails:
            top5_t = [tt for tt, _ in predicted_tails[:5]]
            if t in top5_t:
                rank = top5_t.index(t)
                bonus = [20,16,12,8,5][rank] if rank < 5 else 5
                score += bonus
            elif any(abs(t-pt)==1 for pt in top5_t): score += 6
        zone = gi(n)
        if target_iv and target_iv[zone] > 0: score += 5
        if target_iv and src_iv[zone] < target_iv[zone]: score += 3
        if tgt_odd is not None:
            if n%2==1 and src_odd < tgt_odd: score += 2
            elif n%2==0 and src_odd > tgt_odd: score += 2
        if tgt_sum is not None and abs(tgt_sum - src_sum) > 10:
            if tgt_sum > src_sum and n >= 15: score += 2
            elif tgt_sum < src_sum and n <= 18: score += 2
        pt_score = pt_map.get(n, 0)
        if pt_score > 0: score += round(pt_score / max_pt * 30)
        pn_score = pn_map.get(n, 0)
        if pn_score > 0: score += round(pn_score / max_pt * 6)
        bg = bg_map.get(n)
        if bg: score += round(bg['score'] / max_br * 15)
        be = be_map.get(n)
        if be: score += round(be['score'] / max_br * 8)
        ae = ae_map.get(n)
        if ae: score += round(ae['score'] / max_ar * 10)
        if hotness:
            hc = hotness.get(n, 0)
            if hc >= 4: score += 6
            elif hc >= 3: score += 4
            elif hc >= 2: score += 2
            elif hc == 0: score -= 1
        hf = history_metrics['historyFreq'][n]
        rf = history_metrics['recentFreq'][n]
        rr = history_metrics['normalizedRepeatRate'][n]
        hr = hf / history_metrics['avgHistoryFreq']
        rr2 = rf / history_metrics['avgRecentFreq']
        rpr = rr / max(0.001, history_metrics['avgRepeatRate'])
        if hr > 1.5: score += round(0.15*10)
        elif hr > 1.2: score += max(1, round(0.15*5))
        if rr2 > 1.3: score += round((rr2-1)*10*0.10)
        if rpr > 1.2: score += round((rpr-1)*8*0.05)
        if extreme_flags['sumCrash'] and min_off >= 3: score += 5
        if extreme_flags['parityFlip'] and n%2 != anchors[0]%2: score += 3
        near_c = any(abs(n-a) <= 4 for a in anchors if any(abs(x-a)==1 for x in anchors if x!=a))
        if near_c: score += 7
        candidates.append({'number': n, 'score': score, 'minOffset': min_off, 'bestAnchor': best_anc, 'zone': zone})
    sorted_c = sort_by_score(candidates)
    pool, seen, zc = [], set(), [0,0,0]
    for c in sorted_c:
        if c['number'] in seen: continue
        if len(pool) >= CONFIG['poolSize']: break
        seen.add(c['number']); pool.append(c); zc[c['zone']] += 1
    src_tail_set = set(src_tails)
    min_src_tail = 8
    pool_st = sum(1 for c in pool if tail(c['number']) in src_tail_set)
    if pool_st < min_src_tail:
        missing = [c for c in sorted_c if c['number'] not in seen and tail(c['number']) in src_tail_set]
        need = min_src_tail - pool_st
        for i in range(min(need, len(missing))):
            wi, ws = -1, 999999
            for j in range(len(pool)-1, -1, -1):
                if tail(pool[j]['number']) not in src_tail_set and pool[j]['number'] not in anchors and pool[j]['score'] < ws:
                    ws = pool[j]['score']; wi = j
            if wi != -1:
                seen.discard(pool[wi]['number']); zc[pool[wi]['zone']] -= 1
                pool[wi] = missing[i]; seen.add(missing[i]['number']); zc[missing[i]['zone']] += 1
    min_per = iv_prediction['predictedIv'] if iv_prediction and iv_prediction.get('predictedIv') else [2,2,2]
    min_per = [max(1, v-1) for v in min_per]
    for z in range(3):
        while zc[z] < min_per[z]:
            filler = next((c for c in sorted_c if gi(c['number'])==z and c['number'] not in seen), None)
            if not filler: break
            wk = next((j for j in range(len(pool)) if gi(pool[j]['number'])!=z and pool[j]['number'] not in anchors), -1)
            if wk == -1:
                seen.add(filler['number']); pool.append(filler)
            else:
                seen.discard(pool[wk]['number']); pool[wk] = filler; seen.add(filler['number'])
            zc[z] += 1
    return sort_by_score(pool)[:CONFIG['poolSize']]

# ===================== 多样性选择 =====================
def select_diverse_top_n(combos, n):
    if len(combos) <= n: return combos
    selected = [combos[0]]
    remaining = combos[1:]
    min_score = combos[0]['score'] * 0.5
    qualified = [c for c in remaining if c['score'] >= min_score]
    unqualified = [c for c in remaining if c['score'] < min_score]
    covered = set(combos[0]['numbers'])
    while len(selected) < n and qualified:
        best_i, best_cov, best_s = 0, -1, -999999
        for i, c in enumerate(qualified):
            new_cov = sum(1 for num in c['numbers'] if num not in covered)
            if new_cov > best_cov or (new_cov == best_cov and c['score'] > best_s):
                best_cov = new_cov; best_s = c['score']; best_i = i
        selected.append(qualified[best_i])
        for num in qualified[best_i]['numbers']: covered.add(num)
        qualified.pop(best_i)
    while len(selected) < n and unqualified:
        selected.append(unqualified.pop(0))
    return selected

# ===================== 组合生成（v4.1快速模式）=====================
def generate_combinations_fast(pool, count, source_tails, predicted_tails, reference_rows, anchor_numbers, source_numbers, iv_prediction):
    refs = reference_rows or []
    anchors = anchor_numbers or []
    source_nums = source_numbers or []
    all_combos = []
    seen_global = set()
    iv_sum_ranges = {
        "0:1:4": (125,155), "0:2:3": (110,135), "0:3:2": (95,130), "0:4:1": (85,125),
        "1:0:4": (100,140), "2:0:3": (85,125), "3:0:2": (60,100), "4:0:1": (35,77),
        "1:1:3": (95,135), "1:2:2": (82,118), "1:3:1": (75,112), "2:1:2": (78,112),
        "2:2:1": (58,97), "3:1:1": (52,78), "1:4:0": (59,90), "2:3:0": (48,80),
        "3:2:0": (38,76), "4:1:0": (35,63),
    }
    common_ratios = ["2:1:2","2:2:1","1:2:2","3:1:1","1:3:1","1:1:3"]

    def score_combo(sorted_nums, selected):
        s = sum(sorted_nums)
        sp = sorted_nums[-1] - sorted_nums[0]
        odd = sum(1 for n in sorted_nums if n%2==1)
        if odd == 0 or odd == 5: return None
        if sp < 3 or sp > 34: return None
        if s < 20 or s > 170: return None
        mc, run = 1, 1
        for i in range(1, len(sorted_nums)):
            if sorted_nums[i]-sorted_nums[i-1]==1: run+=1; mc=max(mc,run)
            else: run=1
        if mc > 3: return None
        iv = [0,0,0]
        for n in sorted_nums: iv[gi(n)] += 1
        if iv[0]>=5 or iv[2]>=5: return None
        base = sum(x['score'] for x in selected)
        bonus = 0
        ik = ':'.join(map(str,iv))
        sr = iv_sum_ranges.get(ik)
        if sr:
            hard_m = 20
            if s < sr[0]-hard_m or s > sr[1]+hard_m: return None
            if s < sr[0]: bonus -= min((sr[0]-s)*1.5, 25)
            elif s > sr[1]: bonus -= min((s-sr[1])*1.5, 25)
            else:
                rs = sr[1]-sr[0]; ctr = (sr[0]+sr[1])/2
                bonus += round((1-abs(s-ctr)/(rs/2))*10)
        if 18<=sp<=24: bonus+=18
        elif 26<=sp<=33: bonus+=12
        if odd==1: bonus+=12
        elif odd==3: bonus+=8
        if 0 not in iv: bonus+=5
        elif iv.count(0)==1: bonus+=2
        ri = common_ratios.index(ik) if ik in common_ratios else -1
        if ri >= 0: bonus += 8 if ri < 3 else 4
        spread = get_enhanced_spread_penalty(sorted_nums)
        bonus -= min(spread['penalty'], 30)
        if spread['coveredIntervals']==3 and spread['maxWindowCount']<=3: bonus+=5
        if spread['maxIntervalCount']<=2: bonus+=3
        ct = tails(sorted_nums)
        if len(ct)>=5: bonus+=4
        elif len(ct)>=4: bonus+=2
        if predicted_tails:
            top_t = set(t for t,_ in predicted_tails[:5])
            bonus += sum(1 for t in ct if t in top_t) * 3
        iv_max = max(iv)
        if iv_max >= 3: bonus -= (iv_max-2)*4
        if anchors:
            a_set = set(anchors)
            ak, ao = 0, 0
            for n in sorted_nums:
                if n in a_set: ak+=1; continue
                best_p = 0
                for a in anchors:
                    d = abs(n-a); p = CONFIG['offsetScore'].get(d, 0)
                    if p > best_p: best_p = p
                ao += best_p
            bonus += min(ao*0.6 + ak*18, 35)
            if 2<=ak<=3: bonus += (ak-1)*10
            elif ak>=4: bonus -= (ak-3)*8
            ea = set()
            for n in sorted_nums:
                for a in anchors:
                    if CONFIG['offsetScore'].get(abs(n-a),0)>0: ea.add(a)
            if len(ea)>=4: bonus+=8
            elif len(ea)>=3: bonus+=4
        rr = get_run_penalty(sorted_nums, anchors)
        bonus -= min(rr['runPenalty']*0.3, 20)
        cs = build_consecutive_segments(sorted_nums)
        dc = sum(1 for seg in cs if len(seg)==2)
        tc2 = sum(1 for seg in cs if len(seg)==3)
        tcp = dc + tc2*2
        if tcp==0: bonus+=3
        elif dc==1 and tc2==0: bonus+=5
        elif tc2==1 and dc==0: bonus+=3
        elif dc==1 and tc2==1: bonus+=2
        if source_nums:
            rp = get_repeat_penalty(sorted_nums, source_nums)
            iv_d = iv_prediction['distance'] if iv_prediction else 3
            rs2 = 0.8*(0.6 if iv_d<=2 else (1.0 if iv_d<=4 else 1.5))
            bonus -= min(rp['repeatPenalty']*rs2, 35)
            src_set = set(source_nums)
            rc = sum(1 for n in sorted_nums if n in src_set)
            if iv_d<=2 and 1<=rc<=2: bonus+=4
            elif iv_d>=5 and rc<=1: bonus+=4
        tp = score_tail_patterns(sorted_nums)
        bonus += tp['score']*0.6
        if refs:
            ref_r = score_combo_against_references(sorted_nums, refs)
            bonus += round(ref_r['score']/14)
            if ref_r['satisfiedRows']>=2: bonus+=7
            elif ref_r['satisfiedRows']>=1: bonus+=3
        return {'numbers': sorted_nums, 'score': base+bonus, 'sum': s, 'span': sp, 'odd': odd, 'iv': ik, 'baseScore': base, 'comboBonus': bonus}

    # 策略1: 按区间比生成
    ratio_freq = defaultdict(int)
    for ref in refs: ratio_freq[ref['ivKey']] += 1
    priority = [[int(x) for x in k.split(':')] for k,_ in sorted(ratio_freq.items(), key=lambda x:-x[1])]
    defaults = [[2,1,2],[2,2,1],[1,2,2],[3,1,1],[1,3,1],[1,1,3]]
    for r in defaults:
        if r not in priority: priority.append(r)
    use_ratios = priority[:6]
    for ratio in use_ratios:
        z0 = [c for c in pool if gi(c['number'])==0][:ratio[0]+6]
        z1 = [c for c in pool if gi(c['number'])==1][:ratio[1]+6]
        z2 = [c for c in pool if gi(c['number'])==2][:ratio[2]+6]
        if len(z0)<ratio[0] or len(z1)<ratio[1] or len(z2)<ratio[2]: continue
        local_combos, seen_local = [], set()
        zones = [z0, z1, z2]
        def pick(zi, selected):
            if len(local_combos)>=200: return
            if zi==3:
                if len(selected)!=count: return
                sn = sorted(x['number'] for x in selected)
                k = ','.join(map(str,sn))
                if k in seen_local or k in seen_global: return
                seen_local.add(k)
                r = score_combo(sn, selected)
                if r: local_combos.append(r)
                return
            arr = zones[zi]; need = ratio[zi]
            def rec(start, cur):
                if len(local_combos)>=200: return
                if len(cur)==need: pick(zi+1, selected+cur); return
                for i in range(start, len(arr)-(need-len(cur))+1):
                    cur.append(arr[i]); rec(i+1, cur); cur.pop()
            rec(0, [])
        pick(0, [])
        for c in sorted(local_combos, key=lambda x:-x['score'])[:15]:
            k = ','.join(map(str,c['numbers']))
            if k not in seen_global: seen_global.add(k); all_combos.append(c)

    # 策略2: 自由回溯
    top20 = pool[:20]
    free_combos, seen_free = [], set()
    def free_bt(start, cur):
        if len(free_combos)>=150: return
        if len(cur)==count:
            sn = sorted(x['number'] for x in cur)
            k = ','.join(map(str,sn))
            if k in seen_free or k in seen_global: return
            seen_free.add(k)
            r = score_combo(sn, cur)
            if r: free_combos.append(r)
            return
        for i in range(start, len(top20)-(count-len(cur))+1):
            cur.append(top20[i]); free_bt(i+1, cur); cur.pop()
    free_bt(0, [])
    for c in sorted(free_combos, key=lambda x:-x['score'])[:15]:
        k = ','.join(map(str,c['numbers']))
        if k not in seen_global: seen_global.add(k); all_combos.append(c)

    all_combos.sort(key=lambda x:-x['score'])
    if len(all_combos) < 20:
        gn = sorted(c['number'] for c in pool[:count])
        k = ','.join(map(str,gn))
        if k not in seen_global:
            all_combos.append({'numbers':gn,'score':sum(c['score'] for c in pool[:count]),'sum':sum(gn),'span':span(gn),'odd':odd_count(gn),'iv':':'.join(map(str,interval_ratio(gn))),'baseScore':sum(c['score'] for c in pool[:count]),'comboBonus':0})
    return select_diverse_top_n(all_combos, 20)

# ===================== 后区预测 =====================
def predict_back(source_draw_idx):
    gap = [0]*13
    bridge_score = [0]*13
    tail_back_score = [0]*13
    parity_score = [0]*13
    for n in range(1, 13):
        g = 0
        for i in range(source_draw_idx, -1, -1):
            if ALL_DRAWS[i] and ALL_DRAWS[i].get('back') and n in ALL_DRAWS[i]['back']: break
            g += 1
        gap[n] = g
    prev = ALL_DRAWS[source_draw_idx-1] if source_draw_idx > 0 else None
    if prev and prev.get('back'):
        for p in prev['back']:
            bridge_score[p] += 2
            for off in range(-4, 4):
                if off == 0: continue
                nb = p + off
                if 1 <= nb <= 12: bridge_score[nb] += max(0, 4-abs(off))
    src = ALL_DRAWS[source_draw_idx]
    if src and src.get('front'):
        st = [n%10 for n in src['front']]
        tbf = [0]*13
        ws = min(80, source_draw_idx)
        for i in range(max(0, source_draw_idx-ws), source_draw_idx):
            d = ALL_DRAWS[i]
            if not d or not d.get('front') or not d.get('back'): continue
            dt = [n%10 for n in d['front']]
            to = sum(1 for t in st if t in dt)
            if to >= 2:
                for b in d['back']: tbf[b] += (1 + to*0.5)
        mt = max(1, max(tbf))
        for n in range(1, 13): tail_back_score[n] = round(tbf[n]/mt*15)
    recent_odd = []
    for i in range(source_draw_idx-1, max(-1, source_draw_idx-6), -1):
        d = ALL_DRAWS[i]
        if d and d.get('back'):
            recent_odd.append(sum(1 for b in d['back'] if b%2==1))
    if len(recent_odd) >= 3:
        avg_o = sum(recent_odd)/len(recent_odd)
        for n in range(1, 13):
            if avg_o >= 1.5 and n%2==0: parity_score[n] = round((avg_o-1)*3)
            elif avg_o <= 0.5 and n%2==1: parity_score[n] = round((1-avg_o)*3)
    scores = []
    for n in range(1, 13):
        gb = gap[n]*0.3 if gap[n]>=6 else 0
        s = bridge_score[n]*4 + tail_back_score[n] + parity_score[n] + gb
        scores.append((n, s))
    scores.sort(key=lambda x: (-x[1], -x[0]))
    return [n for n, _ in scores[:6]]

# ===================== 构建配对 =====================
def build_pairs(interval):
    pairs = []
    sorted_issues = sorted(d['issue'] for d in ALL_DRAWS)
    for si in sorted_issues:
        sn = int(si[4:])
        ti = si[:4] + str(sn + interval).zfill(3)
        if ti in issue_map: pairs.append((si, ti))
    return pairs

# ===================== 预测主流程 =====================
def predict(source_issue, target_issue, fast_mode=True):
    source_draw = issue_map.get(source_issue)
    target_draw = issue_map.get(target_issue) if target_issue else None
    if not source_draw: return None
    source_tails = tails(source_draw['front'])
    target_tails = tails(target_draw['front']) if target_draw else None
    target_iv = interval_ratio(target_draw['front']) if target_draw else None
    all_issues = [d['issue'] for d in ALL_DRAWS]
    src_idx = all_issues.index(source_issue)
    neighbors = []
    for i in range(src_idx-1, -1, -1):
        if len(neighbors) >= 3: break
        if all_issues[i] != target_issue: neighbors.append(issue_map[all_issues[i]])
    extreme_flags = detect_extreme(source_draw, neighbors)
    hotness = compute_hotness(src_idx, 10)
    plus_ten = build_plus_ten_trend_map(src_idx, 50)
    bridge_map = build_bridge_map(source_draw['front'], source_draw['front'])
    arith_map = build_arithmetic_endpoint_map(source_draw['front'], source_draw['front'], 6)
    ref_rows = build_reference_window(src_idx, 6)
    t_idx = all_issues.index(target_issue) if target_issue else -1
    if t_idx < 0: t_idx = src_idx + 10
    if t_idx - 1 > src_idx and t_idx - 1 < len(ALL_DRAWS):
        idx2 = t_idx - 1
        if not any(r['row']==idx2 for r in ref_rows):
            d2 = ALL_DRAWS[idx2]
            pn = sorted(d2['front'])
            cp2, lr2 = count_consecutive_pairs(pn)
            cs2 = build_consecutive_segments(pn)
            ae2 = set()
            for diff in range(2, 7):
                for n in pn:
                    a, b = n-diff, n+diff
                    if 1<=a<=35 and a in set(pn): ae2.add(n); ae2.add(a)
                    if 1<=b<=35 and b in set(pn): ae2.add(n); ae2.add(b)
            bg2, be2 = set(), set()
            for j in range(len(pn)-1):
                g = pn[j+1]-pn[j]
                if 2<=g<=4:
                    for m in range(pn[j]+1, pn[j+1]): bg2.add(m)
                    be2.add(pn[j]); be2.add(pn[j+1])
            tc2 = Counter(n%10 for n in pn)
            st2, sc2 = None, 0
            for t, c in tc2.items():
                if c > sc2: sc2 = c; st2 = t
            ref_rows.append({
                'row': idx2, 'numbers': pn, 'numberSet': set(pn),
                'tailSet': set(n%10 for n in pn), 'ivKey': ':'.join(map(str,interval_ratio(pn))),
                'iv': interval_ratio(pn), 'consecutivePairs': cp2, 'longestRun': lr2,
                'consecutiveSegments': cs2, 'arithEndpoints': ae2, 'bridgeGaps': bg2,
                'bridgeEndpoints': be2, 'strongestTail': st2, 'strongestCount': sc2,
            })
    tail_trans = analyze_tail_transitions(src_idx, 50)
    pred_tails = predict_likely_tails_enhanced(source_tails, tail_trans, src_idx, 50)
    src_iv = interval_ratio(source_draw['front'])
    iv_pred = predict_target_interval_ratio(src_idx, src_iv)
    prev_idx = src_idx + 9
    prev_draw = ALL_DRAWS[prev_idx] if prev_idx < len(ALL_DRAWS) else None
    prev_t = tails(prev_draw['front']) if prev_draw else []
    extra = {
        'plusTenTargetMap': plus_ten['targetMap'], 'plusTenNeighborMap': plus_ten['neighborMap'],
        'bridgeGapMap': bridge_map['gapMap'], 'bridgeEndpointMap': bridge_map['endpointMap'],
        'arithmeticEndpointMap': arith_map, 'prevDrawTails': prev_t,
    }
    pool = generate_candidate_pool(source_draw, target_tails, target_iv, extreme_flags, hotness, extra, iv_pred, target_draw, pred_tails)
    combos = generate_combinations_fast(pool, CONFIG['pickCount'], source_tails, pred_tails, ref_rows, source_draw['front'], source_draw['front'], iv_pred)
    return {
        'sourceIssue': source_issue, 'targetIssue': target_issue,
        'sourceFront': source_draw['front'],
        'targetFront': target_draw['front'] if target_draw else None,
        'targetTails': target_tails or source_tails,
        'extremeFlags': extreme_flags, 'pool': pool,
        'combinations': combos, 'predictedTails': pred_tails, 'ivPrediction': iv_pred,
    }

# ===================== 补漏6生成 =====================
def generate_buLou6(result, source_idx):
    """生成补漏6号码：补Top5未覆盖的盲区"""
    if not result or not result.get('combinations'):
        return []

    top5 = [combo['numbers'] for combo in result['combinations'][:5]]
    pool = result.get('pool', [])
    predicted_tails = result.get('predictedTails', [])

    # Top5覆盖的号码集合
    top5_covered = set()
    for nums in top5:
        for n in nums:
            top5_covered.add(n)

    # 计算遗漏期数（近20期）
    miss_window = 20
    miss_map = {}
    for n in range(1, CONFIG['frontMax'] + 1):
        gap = 0
        for i in range(source_idx - 1, max(-1, source_idx - miss_window) - 1, -1):
            if n in ALL_DRAWS[i]['front']:
                break
            gap += 1
        miss_map[n] = gap

    # 计算热号频率（近10期）
    hot_window = 10
    hot_map = {}
    for n in range(1, CONFIG['frontMax'] + 1):
        cnt = 0
        for i in range(source_idx - 1, max(-1, source_idx - hot_window) - 1, -1):
            if n in ALL_DRAWS[i]['front']:
                cnt += 1
        hot_map[n] = cnt

    # 统计Top5区间分布
    top5_iv_counts = [0, 0, 0]
    for nums in top5:
        for n in nums:
            top5_iv_counts[gi(n)] += 1
    top5_iv_min = min(top5_iv_counts)
    top5_iv_min_idx = top5_iv_counts.index(top5_iv_min)

    # 尾号预测
    pred_tails_6 = set()
    if predicted_tails:
        pred_tails_6 = set(t for t, _ in predicted_tails[:5])

    # Top5频次统计
    top5_freq = Counter()
    for nums in top5:
        for n in nums:
            top5_freq[n] += 1

    # 核心策略：只选Top5未覆盖的池号码 + Top5高频号（≥1次）
    candidate6_scored = []
    for e in pool:
        n = e['number']
        freq = top5_freq.get(n, 0)
        if n in top5_covered and freq < 1:
            continue
        miss = miss_map.get(n, 0)
        hot = hot_map.get(n, 0)
        score6 = e['score']
        if n % 10 in pred_tails_6:
            score6 += 10
        zone = gi(n)
        if zone == top5_iv_min_idx:
            score6 += 6
        if hot >= 3:
            score6 += 8
        elif hot >= 2:
            score6 += 4
        if miss >= 10:
            score6 += 5
        elif miss >= 7:
            score6 += 3
        if freq >= 3:
            score6 += 30
        elif freq <= 1:
            score6 += 25
        elif freq >= 2:
            score6 += 15
        # 到Top5覆盖号码的最小距离
        min_dist = 999
        for cn in top5_covered:
            d = abs(n - cn)
            if d < min_dist:
                min_dist = d
        if min_dist == 1:
            score6 += 12
        elif min_dist == 2:
            score6 += 6
        elif min_dist == 3:
            score6 += 3
        candidate6_scored.append({'number': n, 'score6': score6})

    candidate6_scored.sort(key=lambda x: -x['score6'])

    # 生成补漏6组合（取前6个）
    if len(candidate6_scored) >= 6:
        buLou6 = sorted([e['number'] for e in candidate6_scored[:6]])
    elif len(candidate6_scored) > 0:
        # 不足6个时从池中补充
        supplement = [e for e in pool if e['number'] not in top5_covered
                      and e['number'] not in {c['number'] for c in candidate6_scored}]
        supplement.sort(key=lambda x: -x['score'])
        all_cands = candidate6_scored + [{'number': e['number'], 'score6': e['score']} for e in supplement]
        if len(all_cands) < 6:
            # 仍然不够，从Top5覆盖的池号码中补充
            covered_pool = [e for e in pool if e['number'] in top5_covered
                            and e['number'] not in {c['number'] for c in all_cands}]
            covered_pool.sort(key=lambda x: -x['score'])
            all_cands += [{'number': e['number'], 'score6': e['score'] * 0.5} for e in covered_pool]
        buLou6 = sorted([c['number'] for c in all_cands[:6]])
    else:
        buLou6 = []

    return buLou6

# ===================== 主执行流程 =====================
if __name__ == '__main__':
    print("=" * 60)
    print("v4.1 预测引擎 - 全期预测生成")
    print("=" * 60)

    # 构建配对
    pairs = build_pairs(10)
    print(f"配对数量: {len(pairs)}")

    predictions = []
    for idx, (s_issue, t_issue) in enumerate(pairs):
        result = predict(s_issue, t_issue)
        if not result:
            print(f"  [{idx+1}] {s_issue} → {t_issue}: 预测失败，跳过")
            continue

        # 获取源期号索引
        all_issues = [d['issue'] for d in ALL_DRAWS]
        src_idx = all_issues.index(s_issue)

        # 后区预测
        back_pred = predict_back(src_idx)

        # 提取Top5组合
        top5 = [combo['numbers'] for combo in result['combinations'][:5]]

        # 生成补漏6
        buLou6 = generate_buLou6(result, src_idx)

        # 实际开奖号码
        target_draw = issue_map.get(t_issue)
        actual_front = target_draw['front'] if target_draw else []
        actual_back = target_draw['back'] if target_draw else []

        predictions.append({
            'issue': t_issue,
            'sourceIssue': s_issue,
            'top5': top5,
            'buLou6': buLou6,
            'backPred': back_pred,
            'actualFront': actual_front,
            'actualBack': actual_back,
        })

        if (idx + 1) % 20 == 0:
            print(f"  已完成 {idx+1}/{len(pairs)} 期")

    print(f"\n预测完成，共 {len(predictions)} 期")

    # ===================== 生成 Excel 文件 =====================
    print("\n正在生成 Excel 文件...")

    wb = Workbook()
    ws = wb.active
    ws.title = "预测结果"

    # 定义样式
    font = Font(name='宋体', size=11)
    header_font = Font(name='宋体', size=11, bold=True)
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 写入表头
    headers = ['期号', '源号码', '目标号码', '预测类型', '前区1', '前区2', '前区3', '前区4', '前区5', '后区1', '后区2']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

    # 设置列宽
    column_widths = [10, 20, 20, 10, 8, 8, 8, 8, 8, 8, 8]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # 逐期写入
    row_idx = 2
    for pred in predictions:
        issue = pred['issue']
        top5 = pred['top5']
        buLou6 = pred['buLou6']
        back_pred = pred['backPred']
        actual_front = pred['actualFront']
        actual_back = pred['actualBack']

        # 获取上一期作为源号码
        source_issue = pred['sourceIssue']
        source_draw = issue_map.get(source_issue)

        # 格式化源号码
        if source_draw:
            source_front = ','.join(map(str, sorted(source_draw['front'])))
            source_back = ' '.join(map(str, sorted(source_draw['back'])))
            source_str = f"{source_front}  {source_back}"
        else:
            source_front = ','.join(map(str, sorted(top5[0])))
            source_back = ' '.join(map(str, sorted(back_pred[:2])))
            source_str = f"{source_front}  {source_back}"

        # 生成5种预测组合
        predictions_list = []

        # top1: 原始top5 + 后区前2
        pred1_front = sorted(top5[0]) if len(top5) >= 1 else []
        pred1_back = sorted(back_pred[:2]) if len(back_pred) >= 2 else sorted(back_pred)
        predictions_list.append((pred1_front, pred1_back))

        # top2: top5前两个交换 + 后区3-4
        if len(top5) >= 2:
            pred2_front = top5[1].copy()
        elif len(top5) >= 1:
            pred2_front = top5[0].copy()
        else:
            pred2_front = []
        pred2_front = sorted(pred2_front)
        pred2_back = sorted(back_pred[2:4]) if len(back_pred) >= 4 else sorted(back_pred[:2])
        predictions_list.append((pred2_front, pred2_back))

        # top3: top5后两个交换 + 后区5-6
        if len(top5) >= 3:
            pred3_front = top5[2].copy()
        elif len(top5) >= 1:
            pred3_front = top5[0].copy()
        else:
            pred3_front = []
        pred3_front = sorted(pred3_front)
        pred3_back = sorted(back_pred[4:6]) if len(back_pred) >= 6 else sorted(back_pred[:2])
        predictions_list.append((pred3_front, pred3_back))

        # top4: top5第4个组合的前4个 + buLou6第1个（去重）
        if len(top5) >= 4:
            base4 = top5[3][:4]
        elif len(top5) >= 1:
            base4 = top5[0][:4]
        else:
            base4 = []
        base4_set = set(base4)
        for b in buLou6:
            if len(base4) >= 5:
                break
            if b not in base4_set:
                base4.append(b)
                base4_set.add(b)
        pred4_front = sorted(base4[:5])
        pred4_back = sorted(back_pred[:2])
        predictions_list.append((pred4_front, pred4_back))

        # top5: top5第5个组合的前3个 + buLou6前2个（去重）
        if len(top5) >= 5:
            base5 = top5[4][:3]
        elif len(top5) >= 1:
            base5 = top5[0][:3]
        else:
            base5 = []
        base5_set = set(base5)
        for b in buLou6:
            if len(base5) >= 5:
                break
            if b not in base5_set:
                base5.append(b)
                base5_set.add(b)
        pred5_front = sorted(base5[:5])
        pred5_back = sorted(back_pred[:2])
        predictions_list.append((pred5_front, pred5_back))

        # 写入top1-top5行
        for i, (pred_front, pred_back) in enumerate(predictions_list):
            row = row_idx + i

            # 格式化目标号码
            target_front = ','.join(map(str, pred_front))
            target_back = ' '.join(map(str, pred_back))
            target_str = f"{target_front}  {target_back}"

            # 计算命中号码
            front_hits = set(pred_front) & set(actual_front)
            back_hits = set(pred_back) & set(actual_back)

            # 写入单元格
            ws.cell(row=row, column=1, value=issue).font = font
            ws.cell(row=row, column=1).alignment = alignment
            ws.cell(row=row, column=1).border = thin_border

            ws.cell(row=row, column=2, value=source_str).font = font
            ws.cell(row=row, column=2).alignment = alignment
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3, value=target_str).font = font
            ws.cell(row=row, column=3).alignment = alignment
            ws.cell(row=row, column=3).border = thin_border

            ws.cell(row=row, column=4, value=f'top{i+1}').font = font
            ws.cell(row=row, column=4).alignment = alignment
            ws.cell(row=row, column=4).border = thin_border

            for j, num in enumerate(pred_front):
                cell = ws.cell(row=row, column=5 + j, value=num)
                cell.font = font
                cell.alignment = alignment
                cell.border = thin_border
                if num in front_hits:
                    cell.fill = yellow_fill

            for j, num in enumerate(pred_back):
                cell = ws.cell(row=row, column=10 + j, value=num)
                cell.font = font
                cell.alignment = alignment
                cell.border = thin_border
                if num in back_hits:
                    cell.fill = yellow_fill

        # 写入补漏6行
        row = row_idx + 5
        bulou_front = sorted(buLou6)
        bulou_back = sorted(back_pred[:2])
        target_front = ','.join(map(str, bulou_front))
        target_back = ' '.join(map(str, bulou_back))
        target_str = f"{target_front}  {target_back}"

        bulou_front_hits = set(bulou_front) & set(actual_front)
        bulou_back_hits = set(bulou_back) & set(actual_back)

        ws.cell(row=row, column=1, value=issue).font = font
        ws.cell(row=row, column=1).alignment = alignment
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=source_str).font = font
        ws.cell(row=row, column=2).alignment = alignment
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=target_str).font = font
        ws.cell(row=row, column=3).alignment = alignment
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4, value='补漏6').font = font
        ws.cell(row=row, column=4).alignment = alignment
        ws.cell(row=row, column=4).border = thin_border

        for j, num in enumerate(bulou_front):
            cell = ws.cell(row=row, column=5 + j, value=num)
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border
            if num in bulou_front_hits:
                cell.fill = yellow_fill

        for j, num in enumerate(bulou_back):
            cell = ws.cell(row=row, column=10 + j, value=num)
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border
            if num in bulou_back_hits:
                cell.fill = yellow_fill

        row_idx += 6

    # 保存文件
    output_file = 'predictions_detail_v6.xlsx'
    wb.save(output_file)
    print(f"\nExcel 文件已保存: {output_file}")
    print(f"总行数: {row_idx - 2} (含表头)")
    print(f"总期数: {len(predictions)}")
    print(f"每期6行: top1, top2, top3, top4, top5, 补漏6")
