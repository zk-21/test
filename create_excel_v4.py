import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load prediction data
with open('simple_predictions_data.json', 'r', encoding='utf-8') as f:
    predictions = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = "预测结果"

# Define styles
font = Font(name='宋体', size=11)
header_font = Font(name='宋体', size=11, bold=True)
yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers
headers = ['期号', '源号码', '目标号码', '预测类型', '前区1', '前区2', '前区3', '前区4', '前区5', '后区1', '后区2']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment
    cell.border = thin_border

# Set column widths
column_widths = [10, 20, 20, 10, 8, 8, 8, 8, 8, 8, 8]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + i)].width = width

# Process each prediction
row_idx = 2
for pred in predictions:
    issue = pred['issue']
    top5 = pred['top5']
    buLou6 = pred['buLou6']
    backPred = pred['backPred']
    actualFront = pred['actualFront']
    actualBack = pred['actualBack']
    
    # Format source numbers (previous draw numbers - using top5 as source)
    source_front = ','.join(map(str, sorted(top5)))
    source_back = ' '.join(map(str, sorted(backPred[:2])))  # Use first 2 back predictions as source
    source_str = f"{source_front}  {source_back}"
    
    # Format target numbers (prediction results)
    target_front = ','.join(map(str, sorted(top5)))
    target_back = ' '.join(map(str, sorted(backPred[:2])))
    target_str = f"{target_front}  {target_back}"
    
    # Determine hit numbers for highlighting
    front_hits = set(top5) & set(actualFront)
    back_hits = set(backPred[:2]) & set(actualBack)
    
    # Write rows for top1 to top5
    for i in range(5):
        row = row_idx + i
        ws.cell(row=row, column=1, value=issue).font = font
        ws.cell(row=row, column=1).alignment = alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2, value=source_str).font = font
        ws.cell(row=row, column=2).alignment = alignment
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3, value=target_str).font = font
        ws.cell(row=row, column=3).alignment = alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4, value=f'top{i+1}').font = font
        ws.cell(row=row, column=4).alignment = alignment
        ws.cell(row=row, column=4).border = thin_border
        
        # Only fill numbers for top1
        if i == 0:
            for j, num in enumerate(top5):
                cell = ws.cell(row=row, column=5+j, value=num)
                cell.font = font
                cell.alignment = alignment
                cell.border = thin_border
                if num in front_hits:
                    cell.fill = yellow_fill
            
            for j, num in enumerate(backPred[:2]):
                cell = ws.cell(row=row, column=10+j, value=num)
                cell.font = font
                cell.alignment = alignment
                cell.border = thin_border
                if num in back_hits:
                    cell.fill = yellow_fill
        else:
            for col in range(5, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
    
    # Write row for 补漏6
    row = row_idx + 5
    ws.cell(row=row, column=1, value=issue).font = font
    ws.cell(row=row, column=1).alignment = alignment
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=2, value=source_str).font = font
    ws.cell(row=row, column=2).alignment = alignment
    ws.cell(row=row, column=2).border = thin_border
    
    ws.cell(row=row, column=3, value=target_str).font = font
    ws.cell(row=row, column=3).alignment = alignment
    ws.cell(row=row, column=3).border = thin_border
    
    ws.cell(row=row, column=4, value='补漏6').font = font
    ws.cell(row=row, column=4).alignment = alignment
    ws.cell(row=row, column=4).border = thin_border
    
    # Leave prediction columns empty for 补漏6
    for col in range(5, 12):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
    
    row_idx += 6

# Save file
output_file = 'predictions_detail_v4.xlsx'
wb.save(output_file)
print(f"Excel file saved as {output_file}")
print(f"Total rows: {row_idx - 2} (including headers)")
print(f"Total periods: {len(predictions)}")