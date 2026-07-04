import openpyxl
import os

file_path = r"C:\Users\61419\Downloads\2026011-2026012_预测结果.xlsx"
if not os.path.exists(file_path):
    print(f"文件不存在: {file_path}")
    exit(1)

# 读取Excel文件
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("成功读取Excel文件")
    print(f"工作表列表: {wb.sheetnames}")
    
    # 遍历每个工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n工作表 '{sheet_name}' 尺寸: {ws.max_row} 行 × {ws.max_column} 列")
        
        # 读取前10行数据
        print("前10行数据:")
        for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True):
            print(row)
        
        # 读取列名（第一行）
        if ws.max_row > 0:
            headers = [cell.value for cell in ws[1]]
            print(f"\n列名: {headers}")
        
        print("-" * 50)
        
except Exception as e:
    print(f"读取失败: {e}")