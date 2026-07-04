import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 读取预测数据
data_path = os.path.join(os.path.dirname(__file__), 'simple_predictions_data.json')
with open(data_path, 'r', encoding='utf-8') as f:
    predictions = json.load(f)

print(f"加载 {len(predictions)} 期预测数据")

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "预测与命中"

# 定义样式
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')
cell_alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 写入表头
headers = [
    '期号',
    '前区top5_1', '前区top5_2', '前区top5_3', '前区top5_4', '前区top5_5',
    '补漏6_1', '补漏6_2', '补漏6_3', '补漏6_4', '补漏6_5', '补漏6_6',
    '后区预测_1', '后区预测_2', '后区预测_3', '后区预测_4', '后区预测_5', '后区预测_6',
    '实际前区_1', '实际前区_2', '实际前区_3', '实际前区_4', '实际前区_5',
    '实际后区_1', '实际后区_2',
    '前区命中数', '后区命中数'
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 设置列宽
for col in range(1, len(headers) + 1):
    ws.column_dimensions[get_column_letter(col)].width = 12

# 填充数据
for row_idx, pred in enumerate(predictions, 2):
    # 期号
    ws.cell(row=row_idx, column=1, value=pred['issue']).alignment = cell_alignment
    
    # 前区top5
    for i, num in enumerate(pred['top5']):
        cell = ws.cell(row=row_idx, column=2 + i, value=num)
        cell.alignment = cell_alignment
        cell.border = thin_border
    
    # 补漏6
    for i, num in enumerate(pred['buLou6']):
        cell = ws.cell(row=row_idx, column=7 + i, value=num)
        cell.alignment = cell_alignment
        cell.border = thin_border
    
    # 后区预测
    for i, num in enumerate(pred['backPred']):
        cell = ws.cell(row=row_idx, column=13 + i, value=num)
        cell.alignment = cell_alignment
        cell.border = thin_border
    
    # 实际前区
    for i, num in enumerate(pred['actualFront']):
        cell = ws.cell(row=row_idx, column=19 + i, value=num)
        cell.alignment = cell_alignment
        cell.border = thin_border
    
    # 实际后区
    for i, num in enumerate(pred['actualBack']):
        cell = ws.cell(row=row_idx, column=24 + i, value=num)
        cell.alignment = cell_alignment
        cell.border = thin_border
    
    # 计算命中
    actual_front_set = set(pred['actualFront'])
    actual_back_set = set(pred['actualBack'])
    
    # 前区命中：top5和补漏6中与实际前区匹配的号码
    front_hits = 0
    # 检查top5命中
    for num in pred['top5']:
        if num in actual_front_set:
            cell = ws.cell(row=row_idx, column=2 + pred['top5'].index(num))
            cell.fill = yellow_fill
            front_hits += 1
    # 检查补漏6命中
    for num in pred['buLou6']:
        if num in actual_front_set:
            cell = ws.cell(row=row_idx, column=7 + pred['buLou6'].index(num))
            cell.fill = yellow_fill
            front_hits += 1
    
    # 后区命中
    back_hits = 0
    for num in pred['backPred']:
        if num in actual_back_set:
            cell = ws.cell(row=row_idx, column=13 + pred['backPred'].index(num))
            cell.fill = yellow_fill
            back_hits += 1
    
    # 写入命中数
    ws.cell(row=row_idx, column=26, value=front_hits).alignment = cell_alignment
    ws.cell(row=row_idx, column=27, value=back_hits).alignment = cell_alignment

# 冻结首行
ws.freeze_panes = 'A2'

# 添加筛选
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(predictions) + 1}"

# 保存文件
output_path = os.path.join(os.path.dirname(__file__), 'predictions_with_hits.xlsx')
wb.save(output_path)
print(f"Excel文件已保存到 {output_path}")
print(f"共 {len(predictions)} 期数据，命中号码已用黄色背景标记")