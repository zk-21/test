import json
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 加载ALL_DRAWS数据
all_draws_path = os.path.join(os.path.dirname(__file__), 'all_draws.json')
with open(all_draws_path, 'r', encoding='utf-8') as f:
    ALL_DRAWS = json.load(f)
print(f"加载 {len(ALL_DRAWS)} 期数据")

# 构建issue到前区号码的映射
issue_to_front = {}
for draw in ALL_DRAWS:
    issue_to_front[draw['issue']] = draw['front']

# 读取预测数据
data_path = os.path.join(os.path.dirname(__file__), 'simple_predictions_data.json')
with open(data_path, 'r', encoding='utf-8') as f:
    predictions = json.load(f)

print(f"加载 {len(predictions)} 期预测数据")

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "预测明细"

# 定义样式
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')
cell_alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 写入表头 - 按用户要求调整
headers = [
    '期号',
    '源前区1', '源前区2', '源前区3', '源前区4', '源前区5',
    '前区1', '前区2', '前区3', '前区4', '前区5',
    '后区1', '后区2',
    '预测类型'
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 设置列宽
for col in range(1, len(headers) + 1):
    ws.column_dimensions[get_column_letter(col)].width = 12

# 填充数据
row_idx = 2
for pred in predictions:
    issue = pred['issue']
    
    # 源期号计算（前10期）
    try:
        issue_num = int(issue)
        source_issue = str(issue_num - 10) if issue_num > 10 else issue
    except:
        source_issue = issue
    
    # 获取源行前区号码
    source_front = issue_to_front.get(source_issue, [0,0,0,0,0])
    if len(source_front) < 5:
        source_front = [0,0,0,0,0]
    
    # 目标行前区号码
    target_front = pred['actualFront']
    # 目标行后区号码
    target_back = pred['actualBack']
    
    # 预测类型列表（top1-5和补漏6）
    prediction_types = []
    # top1-top5
    for i, num in enumerate(pred['top5']):
        prediction_types.append(('top' + str(i+1), num))
    # 补漏6（6个号码作为一行？用户说"补漏6策略生成的第六行号码"，这里我们显示补漏6的6个号码）
    # 但用户说"第六行号码"，可能是指补漏6的第6个号码？我们显示整个补漏6号码串
    buLou6_str = ','.join(map(str, pred['buLou6']))
    prediction_types.append(('补漏6', buLou6_str))
    
    # 为每个预测类型创建一行
    for pred_type, pred_num in prediction_types:
        # 期号
        ws.cell(row=row_idx, column=1, value=issue).alignment = cell_alignment
        
        # 源前区1-5
        for i, num in enumerate(source_front):
            cell = ws.cell(row=row_idx, column=2 + i, value=num)
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        # 目标前区1-5
        for i, num in enumerate(target_front):
            cell = ws.cell(row=row_idx, column=7 + i, value=num)
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        # 目标后区1-2
        for i, num in enumerate(target_back):
            cell = ws.cell(row=row_idx, column=12 + i, value=num)
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        # 预测类型（包含具体号码）
        if pred_type.startswith('top'):
            pred_label = f"{pred_type}:{pred_num}"
        else:
            pred_label = f"{pred_type}:{pred_num}"
        
        ws.cell(row=row_idx, column=14, value=pred_label).alignment = cell_alignment
        
        row_idx += 1

# 冻结首行
ws.freeze_panes = 'A2'

# 添加筛选
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"

# 保存文件
output_path = os.path.join(os.path.dirname(__file__), 'predictions_detail_v3.xlsx')
wb.save(output_path)
print(f"Excel文件已保存到 {output_path}")
print(f"共 {row_idx - 2} 行数据（{len(predictions)} 期 × 6 个预测类型）")
print("列结构：期号, 源前区1-5, 前区1-5, 后区1-2, 预测类型")