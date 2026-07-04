from openpyxl import load_workbook

# Check the latest generated file
generated_path = "c:/Users/61419/Downloads/selectBall-main/predictions_detail_v6.xlsx"
wb = load_workbook(generated_path, data_only=True)
ws = wb.active
print(f"=== Latest Generated File Info ===")
print(f"Sheet name: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")

print("\nFirst 12 rows (2 periods):")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12, max_col=ws.max_column, values_only=True)):
    print(f"Row {i+1}: {row}")

print("\nColumn headers:")
headers = [cell.value for cell in ws[1]]
print(headers)

# Check if each row has different prediction numbers
print("\n=== Checking prediction numbers for period 2025111 ===")
for row in ws.iter_rows(min_row=2, max_row=7, max_col=ws.max_column, values_only=True):
    if row[0] == '2025111':
        print(f"Issue: {row[0]}, Type: {row[3]}, Front: {row[4:9]}, Back: {row[9:11]}")

# Check if source and target numbers are different
print("\n=== Checking source vs target numbers ===")
for row in ws.iter_rows(min_row=2, max_row=7, max_col=3, values_only=True):
    if row[0] and row[1] and row[2]:
        print(f"Issue: {row[0]}, Source: {row[1]}, Target: {row[2]}")
        if row[1] == row[2]:
            print("  WARNING: Source and target are the same!")
        else:
            print("  OK: Source and target are different")

# Check if yellow highlighting exists
print("\n=== Checking for yellow highlighting ===")
yellow_count = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        if cell.fill.start_color and cell.fill.start_color.rgb == 'FFFFFF00':
            yellow_count += 1
            if yellow_count <= 5:  # Show first 5 examples
                print(f"Yellow cell at {cell.coordinate}: value='{cell.value}'")
print(f"Total yellow cells: {yellow_count}")