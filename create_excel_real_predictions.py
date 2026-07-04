# -*- coding: utf-8 -*-
import json
import sys
import io

# Fix Windows encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load real prediction data
with open('real_predictions_2025101_2025111.json', 'r', encoding='utf-8') as f:
    predictions = json.load(f)

print(f"Loaded {len(predictions)} predictions")

# Load all_draws.json for source back numbers
with open('all_draws.json', 'r', encoding='utf-8') as f:
    all_draws = json.load(f)

draw_map = {draw['issue']: draw for draw in all_draws}

wb = Workbook()
ws = wb.active
ws.title = "Predictions"

# Styles
font = Font(name='Arial', size=11)
header_font = Font(name='Arial', size=11, bold=True)
yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ['期号', '源号码', '目标号码', '预测类型', '前区1', '前区2', '前区3', '前区4', '前区5', '后区1', '后区2']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment
    cell.border = thin_border

# Column widths
column_widths = [12, 22, 22, 10, 8, 8, 8, 8, 8, 8, 8]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + i)].width = width

# Process each prediction
row_idx = 2
for pred in predictions:
    tgt_issue = pred['tgtIssue']
    src_front = pred['srcFront']
    tgt_front = pred['tgtFront']
    tgt_back = pred['tgtBack']
    top5 = pred['top5']
    buLou6 = pred['buLou6']
    back_pred = pred['backPred']
    
    # Get source back numbers from all_draws
    src_draw = draw_map.get(pred['srcIssue'])
    src_back = src_draw['back'] if src_draw else []
    
    # Format source numbers (front + back)
    source_front_str = ','.join(map(str, sorted(src_front)))
    source_back_str = ','.join(map(str, sorted(src_back)))
    source_str = f"{source_front_str}  {source_back_str}"
    
    # Format target numbers
    if tgt_front and tgt_back:
        target_front_str = ','.join(map(str, sorted(tgt_front)))
        target_back_str = ','.join(map(str, sorted(tgt_back)))
        target_str = f"{target_front_str}  {target_back_str}"
    else:
        target_str = ""
    
    # Target sets for hit detection
    target_front_set = set(tgt_front) if tgt_front else set()
    target_back_set = set(tgt_back) if tgt_back else set()
    
    # Write Top1 to Top5
    for i in range(5):
        combo = top5[i]
        pred_front = combo['numbers']  # 5 numbers
        pred_back = combo['back']      # each combo has its own back pair
        
        pred_type = f"Top{i+1}"
        
        ws.cell(row=row_idx, column=1, value=tgt_issue)
        ws.cell(row=row_idx, column=2, value=source_str)
        ws.cell(row=row_idx, column=3, value=target_str)
        ws.cell(row=row_idx, column=4, value=pred_type)
        
        # Front zone (5 numbers)
        for j in range(5):
            cell = ws.cell(row=row_idx, column=5+j, value=pred_front[j])
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border
            if pred_front[j] in target_front_set:
                cell.fill = yellow_fill
        
        # Back zone (2 numbers)
        for j in range(2):
            cell = ws.cell(row=row_idx, column=10+j, value=pred_back[j])
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border
            if pred_back[j] in target_back_set:
                cell.fill = yellow_fill
        
        # Style info columns
        for col in range(1, 5):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border
        
        row_idx += 1
    
    # Write buLou6
    buLou_front = buLou6['numbers'][:5]  # first 5 of 6
    buLou_back = buLou6['back']          # buLou6 has its own back pair
    
    ws.cell(row=row_idx, column=1, value=tgt_issue)
    ws.cell(row=row_idx, column=2, value=source_str)
    ws.cell(row=row_idx, column=3, value=target_str)
    ws.cell(row=row_idx, column=4, value='补漏6')
    
    # Front zone
    for j in range(5):
        cell = ws.cell(row=row_idx, column=5+j, value=buLou_front[j])
        cell.font = font
        cell.alignment = alignment
        cell.border = thin_border
        if buLou_front[j] in target_front_set:
            cell.fill = yellow_fill
    
    # Back zone
    for j in range(2):
        cell = ws.cell(row=row_idx, column=10+j, value=buLou_back[j])
        cell.font = font
        cell.alignment = alignment
        cell.border = thin_border
        if buLou_back[j] in target_back_set:
            cell.fill = yellow_fill
    
    # Style info columns
    for col in range(1, 5):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = font
        cell.alignment = alignment
        cell.border = thin_border
    
    row_idx += 1

# Save
output_file = 'real_predictions_2025101_2025111.xlsx'
wb.save(output_file)

print(f"Excel file saved: {output_file}")
print(f"  {len(predictions)} predictions x 6 rows = {len(predictions) * 6} total rows")

# Show sample
if predictions:
    sample = predictions[0]
    print(f"\nSample ({sample['srcIssue']} -> {sample['tgtIssue']}):")
    print(f"  Source: {sample['srcFront']}")
    print(f"  Target: {sample['tgtFront']}")
    print(f"  Top1: {sample['top5'][0]['numbers']}")
    print(f"  Top2: {sample['top5'][1]['numbers']}")
    print(f"  Top3: {sample['top5'][2]['numbers']}")
    print(f"  Top4: {sample['top5'][3]['numbers']}")
    print(f"  Top5: {sample['top5'][4]['numbers']}")
    print(f"  BuLou6: {sample['buLou6']['numbers']}")
    print(f"  Back: {sample['backPred'][:2]}")
