import openpyxl
import json

with open('all_draws.json', 'r', encoding='utf-8') as f:
    ALL_DRAWS = json.load(f)
issue_map = {d['issue']: d for d in ALL_DRAWS}

wb = openpyxl.load_workbook('predictions_detail_v6.xlsx')
ws = wb.active

# 按期号分组
periods = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    issue = row[0]
    pred_type = row[3]
    front = set([row[4], row[5], row[6], row[7], row[8]])
    back = set([row[9], row[10]])
    if issue not in periods:
        periods[issue] = {}
    periods[issue][pred_type] = {'front': front, 'back': back}

# 统计各种覆盖率
total = 0
# Top1+补漏6联合前区覆盖
combined_ge1 = 0
combined_ge2 = 0
combined_ge3 = 0
# Top5全部组合联合覆盖（5组×5球=25球，有重复）
top5_all_ge1 = 0
top5_all_ge2 = 0
# 候选池覆盖率（top1~top5 + 补漏6 = 11组号码的并集）
all_pred_ge1 = 0
all_pred_ge2 = 0
all_pred_ge3 = 0
all_pred_ge4 = 0
all_pred_ge5 = 0

for issue, preds in periods.items():
    actual = issue_map.get(issue)
    if not actual:
        continue
    actual_front = set(actual['front'])
    actual_back = set(actual['back'])
    total += 1

    # Top1+补漏6联合覆盖
    combined_front = set()
    if 'top1' in preds:
        combined_front |= preds['top1']['front']
    if '补漏6' in preds:
        combined_front |= preds['补漏6']['front']
    hit_c = len(actual_front & combined_front)
    if hit_c >= 1: combined_ge1 += 1
    if hit_c >= 2: combined_ge2 += 1
    if hit_c >= 3: combined_ge3 += 1

    # Top5全部组合联合覆盖
    top5_all_front = set()
    for pt in ['top1', 'top2', 'top3', 'top4', 'top5']:
        if pt in preds:
            top5_all_front |= preds[pt]['front']
    hit_5 = len(actual_front & top5_all_front)
    if hit_5 >= 1: top5_all_ge1 += 1
    if hit_5 >= 2: top5_all_ge2 += 1

    # 全部预测联合覆盖（Top1~Top5 + 补漏6）
    all_front = set()
    for pt in ['top1', 'top2', 'top3', 'top4', 'top5', '补漏6']:
        if pt in preds:
            all_front |= preds[pt]['front']
    hit_all = len(actual_front & all_front)
    if hit_all >= 1: all_pred_ge1 += 1
    if hit_all >= 2: all_pred_ge2 += 1
    if hit_all >= 3: all_pred_ge3 += 1
    if hit_all >= 4: all_pred_ge4 += 1
    if hit_all >= 5: all_pred_ge5 += 1

print(f"回测期数: {total}")
print()

print("=== Top1+补漏6 联合前区覆盖（10球 vs 实际5球） ===")
print(f"  命中>=1个: {combined_ge1}/{total} ({combined_ge1/total*100:.1f}%)")
print(f"  命中>=2个: {combined_ge2}/{total} ({combined_ge2/total*100:.1f}%)")
print(f"  命中>=3个: {combined_ge3}/{total} ({combined_ge3/total*100:.1f}%)")

print()
print("=== Top5全部组合联合前区覆盖（约25球去重 vs 实际5球） ===")
print(f"  命中>=1个: {top5_all_ge1}/{total} ({top5_all_ge1/total*100:.1f}%)")
print(f"  命中>=2个: {top5_all_ge2}/{total} ({top5_all_ge2/total*100:.1f}%)")

print()
print("=== Top1~Top5+补漏6 全部联合前区覆盖（约30球去重 vs 实际5球） ===")
print(f"  命中>=1个: {all_pred_ge1}/{total} ({all_pred_ge1/total*100:.1f}%)")
print(f"  命中>=2个: {all_pred_ge2}/{total} ({all_pred_ge2/total*100:.1f}%)")
print(f"  命中>=3个: {all_pred_ge3}/{total} ({all_pred_ge3/total*100:.1f}%)")
print(f"  命中>=4个: {all_pred_ge4}/{total} ({all_pred_ge4/total*100:.1f}%)")
print(f"  命中=5个:  {all_pred_ge5}/{total} ({all_pred_ge5/total*100:.1f}%)")
