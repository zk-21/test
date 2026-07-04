import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 读取预测数据
data_path = os.path.join(os.path.dirname(__file__), 'simple_predictions_data.json')
with open(data_path, 'r', encoding='utf-8') as f:
    predictions = json.load(f)

print(f"加载 {len(predictions)} 期预测数据")

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "预测明细"

# 定义样式
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')
cell_alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 写入表头
headers = [
    '期号', '源号码', '目标号码',
    '前区1', '前区2', '前区3', '前区4', '前区5',
    '后区1', '后区2',
    '预测类型', '预测号码', '是否命中'
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 设置列宽
for col in range(1, len(headers) + 1):
    ws.column_dimensions[get_column_letter(col)].width = 12

# 填充数据
row_idx = 2
for pred in predictions:
    issue = pred['issue']
    # 源号码：前10期的期号（简化：使用issue-10）
    # 这里需要计算源期号，但我们的数据中没有存储源期号
    # 暂时使用目标期号减去10作为示例
    try:
        issue_num = int(issue)
        source_issue = str(issue_num - 10) if issue_num > 10 else issue
    except:
        source_issue = issue
    target_issue = issue
    
    # 实际开奖号码
    actual_front = pred['actualFront']
    actual_back = pred['actualBack']
    
    # 预测类型列表
    prediction_types = []
    # top1-top5
    for i, num in enumerate(pred['top5']):
        prediction_types.append(('top' + str(i+1), num))
    # 补漏1-补漏6
    for i, num in enumerate(pred['buLou6']):
        prediction_types.append(('补漏' + str(i+1), num))
    
    # 为每个预测类型创建一行
    for pred_type, pred_num in prediction_types:
        # 期号
        ws.cell(row=row_idx, column=1, value=issue).alignment = cell_alignment
        # 源号码
        ws.cell(row=row_idx, column=2, value=source_issue).alignment = cell_alignment
        # 目标号码
        ws.cell(row=row_idx, column=3, value=target_issue).alignment = cell_alignment
        
        # 实际前区1-5
        for i, num in enumerate(actual_front):
            cell = ws.cell(row=row_idx, column=4 + i, value=num)
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        # 实际后区1-2
        for i, num in enumerate(actual_back):
            cell = ws.cell(row=row_idx, column=9 + i, value=num)
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        # 预测类型
        ws.cell(row=row_idx, column=11, value=pred_type).alignment = cell_alignment
        # 预测号码
        cell = ws.cell(row=row_idx, column=12, value=pred_num)
        cell.alignment = cell_alignment
        cell.border = thin_border
        
        # 判断是否命中
        actual_front_set = set(actual_front)
        actual_back_set = set(actual_back)
        
        is_hit = False
        if pred_type.startswith('top') or pred_type.startswith('补漏'):
            # 前区预测
            if pred_num in actual_front_set:
                is_hit = True
                cell.fill = yellow_fill
        
        # 是否命中
        ws.cell(row=row_idx, column=13, value='是' if is_hit else '否').alignment = cell_alignment
        
        row_idx += 1

# 冻结首行
ws.freeze_panes = 'A2'

# 添加筛选
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"

# 保存文件
output_path = os.path.join(os.path.dirname(__file__), 'predictions_detail.xlsx')
wb.save(output_path)
print(f"Excel文件已保存到 {output_path}")
print(f"共 {row_idx - 2} 行数据（{len(predictions)} 期 × 11 个预测号码）")
print("命中号码已用黄色背景标记")