import json

with open('all_draws.json', 'r', encoding='utf-8') as f:
    ALL_DRAWS = json.load(f)

issue_map = {d['issue']: d for d in ALL_DRAWS}

# 2026072期
draw_72 = issue_map.get('2026072')
if draw_72:
    print(f"2026072期: 前区 {draw_72['front']}, 后区 {draw_72['back']}")

# 2026062期（源期号，间隔10）
draw_62 = issue_map.get('2026062')
if draw_62:
    print(f"2026062期: 前区 {draw_62['front']}, 后区 {draw_62['back']}")

# 检查predictions_detail_v6.xlsx中2026072期的数据
import openpyxl
wb = openpyxl.load_workbook('predictions_detail_v6.xlsx')
ws = wb.active

print("\n=== Excel中2026072期的数据 ===")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] == '2026072':
        print(f"类型: {row[3]}, 前区: {row[4:9]}, 后区: {row[9:11]}")
