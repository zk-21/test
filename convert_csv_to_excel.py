import pandas as pd
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def convert_csv_to_excel(csv_path, excel_path):
    """
    将CSV文件转换为Excel文件，并标记命中的号码为黄色
    """
    # 读取CSV文件
    df = pd.read_csv(csv_path)
    
    # 确保列顺序正确
    expected_columns = ['期号', '源号码', '目标号码', '预测类型', '前区1', '前区2', '前区3', '前区4', '前区5', '后区1', '后区2']
    
    # 检查列是否存在
    missing_columns = [col for col in expected_columns if col not in df.columns]
    if missing_columns:
        print(f"警告: CSV文件缺少以下列: {missing_columns}")
        for col in missing_columns:
            df[col] = ''
    
    # 重新排列列顺序
    df = df[expected_columns]
    
    # 写入Excel文件
    df.to_excel(excel_path, index=False, engine='openpyxl')
    
    # 加载工作簿，添加黄色标记
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 黄色填充样式
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # 遍历每一行（从第2行开始，第1行是表头）
    for row_idx in range(2, ws.max_row + 1):
        # 获取目标号码
        target_cell = ws.cell(row=row_idx, column=3)  # 第3列是"目标号码"
        target_str = str(target_cell.value) if target_cell.value else ''
        
        # 解析目标号码（可能是逗号分隔或空格分隔）
        target_nums = set()
        if target_str:
            # 处理各种分隔符
            for num_str in target_str.replace(',', ' ').split():
                try:
                    target_nums.add(int(num_str.strip()))
                except ValueError:
                    pass
        
        # 检查前区1-5和后区1-2（列索引5-11，即第6列到第11列）
        for col_idx in range(5, 12):  # 前区1到后区2
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                cell_value = int(cell.value) if cell.value else None
                if cell_value is not None and cell_value in target_nums:
                    cell.fill = yellow_fill
            except (ValueError, TypeError):
                pass
    
    # 保存修改后的文件
    wb.save(excel_path)
    
    print(f"已成功将CSV转换为Excel: {excel_path}")
    print(f"总行数: {len(df)}")
    print(f"列数: {len(df.columns)}")
    print(f"已添加黄色标记命中的号码")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: python convert_csv_to_excel.py <输入CSV文件> <输出Excel文件>")
        print("示例: python convert_csv_to_excel.py predictions_detail_v6.csv predictions_detail_v6.xlsx")
        sys.exit(1)
    
    csv_path = sys.argv[1]
    excel_path = sys.argv[2]
    
    if not os.path.exists(csv_path):
        print(f"错误: CSV文件不存在: {csv_path}")
        sys.exit(1)
    
    convert_csv_to_excel(csv_path, excel_path)